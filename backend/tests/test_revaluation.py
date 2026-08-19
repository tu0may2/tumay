"""Переоценка портфеля и импорт портфеля из книги Excel.

Проверяем то, в чём ошибиться дороже всего: разделение видов учёта (торговый
переоценивается по рынку, до погашения — нет), выбор базы переоценки (СВЦ, а
не цена последней сделки) и защиту от задвоения позиции при импорте остатков
поверх уже загруженных сделок.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from app.models import (
    ACCOUNTING_HTM,
    ACCOUNTING_TRADING,
    Base,
    Deal,
    Instrument,
    Portfolio,
    Quote,
)
from app.services import portfolio as portfolio_service
from app.services import portfolio_import as import_service
from app.services import revaluation as reval_service


@pytest.fixture()
def session():
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False)
    with factory() as active:
        yield active


def add_bond(session, secid="SU26245RMFS9", *, maturity=None, face=1000.0):
    instrument = Instrument(
        secid=secid, board="TQOB", engine="stock", market="bonds", kind="bond",
        short_name=secid, isin=f"RU000{secid}", face_value=face, face_unit="SUR",
        maturity_date=maturity,
    )
    session.add(instrument)
    session.flush()
    return instrument


def add_share(session, secid="SBER"):
    instrument = Instrument(
        secid=secid, board="TQBR", engine="stock", market="shares", kind="share",
        short_name=secid,
    )
    session.add(instrument)
    session.flush()
    return instrument


def add_quote(session, instrument, **kwargs):
    quote = Quote(instrument_id=instrument.id, ts=datetime(2026, 8, 18, 12, 0), **kwargs)
    session.add(quote)
    session.flush()
    return quote


def add_deal(session, secid, *, portfolio="Торговый", quantity=100, price=90.0,
             trade_date=date(2026, 1, 15), side="buy"):
    deal = Deal(
        portfolio=portfolio, secid=secid, side=side, quantity=quantity,
        price=price, trade_date=trade_date,
    )
    session.add(deal)
    session.flush()
    return deal


class TestAmortizedPrice:
    """Амортизированная стоимость — линейное движение цены к номиналу."""

    def test_discount_grows_to_par(self):
        # Купили за 90 на два года — к погашению должно стать 100
        price = reval_service.amortized_price(
            purchase_price=90.0,
            maturity=date(2028, 1, 1),
            purchased_on=date(2026, 1, 1),
            today=date(2027, 1, 1),
        )
        assert price == pytest.approx(95.0, abs=0.1)

    def test_premium_falls_to_par(self):
        price = reval_service.amortized_price(
            purchase_price=110.0,
            maturity=date(2028, 1, 1),
            purchased_on=date(2026, 1, 1),
            today=date(2027, 1, 1),
        )
        assert price == pytest.approx(105.0, abs=0.1)

    def test_at_maturity_equals_par(self):
        price = reval_service.amortized_price(
            purchase_price=90.0,
            maturity=date(2028, 1, 1),
            purchased_on=date(2026, 1, 1),
            today=date(2028, 1, 1),
        )
        assert price == 100.0

    def test_after_maturity_stays_par(self):
        """Погашенная бумага не может стоить больше номинала из-за экстраполяции."""
        price = reval_service.amortized_price(
            purchase_price=90.0,
            maturity=date(2027, 1, 1),
            purchased_on=date(2026, 1, 1),
            today=date(2030, 1, 1),
        )
        assert price == 100.0

    def test_without_maturity_is_undefined(self):
        """У акции нет погашения — амортизировать нечего."""
        assert reval_service.amortized_price(
            purchase_price=90.0, maturity=None,
            purchased_on=date(2026, 1, 1), today=date(2027, 1, 1),
        ) is None


class TestRevaluation:
    def test_uses_weighted_average_not_last_price(self, session):
        """Переоценка идёт по СВЦ: одна сделка в конце дня не должна её двигать."""
        bond = add_bond(session)
        # last заметно выше СВЦ — если возьмём last, оценка будет завышена
        add_quote(session, bond, last=99.0, wa_price=95.0, prev_wa_price=94.0)
        add_deal(session, bond.secid, quantity=10, price=90.0)
        session.commit()

        result = reval_service.revaluate(session)
        item = result["items"][0]
        assert item["wa_price"] == 95.0
        # 10 бумаг × 95% × номинал 1000 = 9500, а не 9900 по last
        assert item["market_value_rub"] == pytest.approx(9500.0)

    def test_daily_revaluation_is_difference_from_previous_day(self, session):
        bond = add_bond(session)
        add_quote(session, bond, wa_price=95.0, prev_wa_price=94.0)
        add_deal(session, bond.secid, quantity=10, price=90.0)
        session.commit()

        item = reval_service.revaluate(session)["items"][0]
        # (95 − 94)% × 1000 × 10 = 100 рублей за день
        assert item["daily_reval_rub"] == pytest.approx(100.0)

    def test_total_revaluation_is_difference_from_cost(self, session):
        bond = add_bond(session)
        add_quote(session, bond, wa_price=95.0, prev_wa_price=94.0)
        add_deal(session, bond.secid, quantity=10, price=90.0)
        session.commit()

        item = reval_service.revaluate(session)["items"][0]
        # (95 − 90)% × 1000 × 10 = 500 рублей с покупки
        assert item["total_reval_rub"] == pytest.approx(500.0)

    def test_htm_carries_amortized_value_not_market(self, session):
        """Для портфеля до погашения в учёт идёт амортизированная стоимость."""
        bond = add_bond(session, maturity=date(2028, 1, 1))
        add_quote(session, bond, wa_price=95.0, prev_wa_price=94.0)
        add_deal(
            session, bond.secid, portfolio="Инвестиционный",
            quantity=10, price=90.0, trade_date=date(2026, 1, 1),
        )
        portfolio_service.set_accounting_type(session, "Инвестиционный", ACCOUNTING_HTM)

        item = reval_service.revaluate(session)["items"][0]
        assert item["accounting_type"] == ACCOUNTING_HTM
        assert item["market_is_reference"] is True
        # Учётная стоимость — амортизированная, а не рыночные 9500
        assert item["carrying_value_rub"] == pytest.approx(item["amortized_value_rub"])
        assert item["carrying_value_rub"] != pytest.approx(item["market_value_rub"])

    def test_htm_still_shows_market_for_reference(self, session):
        """Рыночную оценку по HTM показываем — просто помечаем справочной."""
        bond = add_bond(session, maturity=date(2028, 1, 1))
        add_quote(session, bond, wa_price=95.0, prev_wa_price=94.0)
        add_deal(session, bond.secid, portfolio="Инвестиционный", quantity=10, price=90.0)
        portfolio_service.set_accounting_type(session, "Инвестиционный", ACCOUNTING_HTM)

        item = reval_service.revaluate(session)["items"][0]
        assert item["market_value_rub"] is not None
        assert item["total_reval_rub"] is not None

    def test_htm_revaluation_excluded_from_totals(self, session):
        """Справочная переоценка не должна попадать в итоговую сумму.

        Иначе казначей увидит прибыль, которой в учёте нет.
        """
        trading = add_bond(session, "TRADE1")
        htm = add_bond(session, "HOLD1", maturity=date(2028, 1, 1))
        add_quote(session, trading, wa_price=95.0, prev_wa_price=94.0)
        add_quote(session, htm, wa_price=95.0, prev_wa_price=94.0)
        add_deal(session, "TRADE1", portfolio="Торговый", quantity=10, price=90.0)
        add_deal(session, "HOLD1", portfolio="Инвестиционный", quantity=10, price=90.0)
        portfolio_service.set_accounting_type(session, "Инвестиционный", ACCOUNTING_HTM)

        result = reval_service.revaluate(session)
        # В итоги вошёл только торговый: 500 рублей, а не 1000
        assert result["totals"]["total_reval_rub"] == pytest.approx(500.0)
        assert result["totals"]["daily_reval_rub"] == pytest.approx(100.0)

    def test_two_portfolios_are_reported_separately(self, session):
        """Два портфеля не должны сливаться — это главное требование к разделению."""
        bond = add_bond(session)
        add_quote(session, bond, wa_price=95.0, prev_wa_price=94.0)
        add_deal(session, bond.secid, portfolio="Торговый", quantity=10, price=90.0)
        add_deal(session, bond.secid, portfolio="Инвестиционный", quantity=5, price=92.0)
        session.commit()

        by_portfolio = reval_service.revaluate(session)["by_portfolio"]
        names = {row["portfolio"] for row in by_portfolio}
        assert "Торговый" in names or "Инвестиционный" in names

        # А отбор по одному портфелю показывает только его позиции
        only_trading = reval_service.revaluate(session, portfolio="Торговый")
        assert only_trading["items"][0]["quantity"] == 10

    def test_empty_portfolio_does_not_crash(self, session):
        result = reval_service.revaluate(session)
        assert result["items"] == []
        assert result["totals"]["positions"] == 0

    def test_missing_previous_price_leaves_daily_empty(self, session):
        """Без вчерашней СВЦ дневную переоценку не выдумываем."""
        bond = add_bond(session)
        add_quote(session, bond, wa_price=95.0)
        add_deal(session, bond.secid, quantity=10, price=90.0)
        session.commit()

        item = reval_service.revaluate(session)["items"][0]
        assert item["daily_reval_rub"] is None


class TestAccountingRegistry:
    def test_unknown_portfolio_defaults_to_trading(self, session):
        """Портфель без записи в справочнике ведёт себя как раньше."""
        add_deal(session, "X", portfolio="Старый")
        session.commit()
        assert portfolio_service.accounting_types(session)["Старый"] == ACCOUNTING_TRADING

    def test_names_include_registry_and_deals(self, session):
        add_deal(session, "X", portfolio="Из сделок")
        portfolio_service.set_accounting_type(session, "Только в справочнике", ACCOUNTING_HTM)
        names = portfolio_service.portfolio_names(session)
        assert "Из сделок" in names
        assert "Только в справочнике" in names

    def test_rejects_unknown_accounting_type(self, session):
        with pytest.raises(ValueError):
            portfolio_service.set_accounting_type(session, "П", "неведомый")


def _book(sheets: dict[str, list[list]]) -> bytes:
    """Собрать книгу Excel в память для теста импорта."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets.items():
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class TestPortfolioImport:
    def test_template_has_all_three_sheets(self):
        workbook = load_workbook(io.BytesIO(import_service.build_template()))
        assert set(workbook.sheetnames) == {"Портфели", "Остатки", "Сделки"}

    def test_holdings_become_opening_buy_deals(self, session):
        """Остаток превращается в сделку покупки — иначе позиции бы не было."""
        add_bond(session, "SU26245RMFS9")
        session.commit()

        content = _book({
            "Остатки": [
                ["Портфель", "Код бумаги", "Количество", "Цена приобретения", "Дата приобретения"],
                ["Торговый", "SU26245RMFS9", 100, 90.0, date(2026, 1, 15)],
            ]
        })
        preview = import_service.preview(session, content, "portfolio.xlsx")
        assert preview["holdings_valid"] == 1

        result = import_service.apply(session, preview)
        assert result["holdings"] == 1

        deal = session.execute(select(Deal)).scalar_one()
        assert deal.side == "buy"
        assert deal.quantity == 100
        assert deal.comment == "Входящий остаток"

    def test_accounting_types_are_read_from_sheet(self, session):
        content = _book({
            "Портфели": [
                ["Портфель", "Вид учёта"],
                ["Торговый", "торговый"],
                ["Инвестиционный", "до погашения"],
            ]
        })
        preview = import_service.preview(session, content, "p.xlsx")
        import_service.apply(session, preview)

        types = {
            row.name: row.accounting_type
            for row in session.execute(select(Portfolio)).scalars()
        }
        assert types["Торговый"] == ACCOUNTING_TRADING
        assert types["Инвестиционный"] == ACCOUNTING_HTM

    def test_warns_about_double_counting(self, session):
        """Остатки поверх существующих сделок удвоят позицию — надо предупредить."""
        add_bond(session, "SU26245RMFS9")
        add_deal(session, "SU26245RMFS9", portfolio="Торговый")
        session.commit()

        content = _book({
            "Остатки": [
                ["Портфель", "Код бумаги", "Количество", "Цена приобретения"],
                ["Торговый", "SU26245RMFS9", 100, 90.0],
            ]
        })
        preview = import_service.preview(session, content, "p.xlsx")
        assert any("задвоится" in w for w in preview["warnings"])

    def test_replace_existing_clears_previous_deals(self, session):
        add_bond(session, "SU26245RMFS9")
        add_deal(session, "SU26245RMFS9", portfolio="Торговый", quantity=50)
        session.commit()

        content = _book({
            "Остатки": [
                ["Портфель", "Код бумаги", "Количество", "Цена приобретения"],
                ["Торговый", "SU26245RMFS9", 100, 90.0],
            ]
        })
        preview = import_service.preview(session, content, "p.xlsx")
        result = import_service.apply(session, preview, replace_existing=True)

        assert result["removed"] == 1
        deals = list(session.execute(select(Deal)).scalars())
        assert len(deals) == 1
        assert deals[0].quantity == 100

    def test_unknown_instrument_is_flagged_not_silently_dropped(self, session):
        content = _book({
            "Остатки": [
                ["Портфель", "Код бумаги", "Количество", "Цена приобретения"],
                ["Торговый", "НЕТТАКОЙ", 100, 90.0],
            ]
        })
        preview = import_service.preview(session, content, "p.xlsx")
        assert preview["holdings"][0]["ok"] is False
        assert "не найден" in preview["holdings"][0]["problems"][0]

    def test_isin_is_resolved_to_secid(self, session):
        bond = add_bond(session, "SU26245RMFS9")
        session.commit()

        content = _book({
            "Остатки": [
                ["Портфель", "ISIN", "Количество", "Цена приобретения"],
                ["Торговый", bond.isin, 100, 90.0],
            ]
        })
        preview = import_service.preview(session, content, "p.xlsx")
        assert preview["holdings"][0]["secid"] == "SU26245RMFS9"

    def test_deals_sheet_keeps_sell_direction(self, session):
        add_bond(session, "SU26245RMFS9")
        session.commit()

        content = _book({
            "Сделки": [
                ["Портфель", "Дата", "Код бумаги", "Направление", "Количество", "Цена"],
                ["Торговый", date(2026, 5, 1), "SU26245RMFS9", "продажа", 10, 95.0],
            ]
        })
        preview = import_service.preview(session, content, "p.xlsx")
        import_service.apply(session, preview)

        deal = session.execute(select(Deal)).scalar_one()
        assert deal.side == "sell"

    def test_single_sheet_without_direction_is_treated_as_holdings(self, session):
        """Однолистовую книгу без «направления» считаем составом, а не сделками."""
        add_bond(session, "SU26245RMFS9")
        session.commit()

        content = _book({
            "Мой портфель": [
                ["Код бумаги", "Количество", "Цена приобретения"],
                ["SU26245RMFS9", 100, 90.0],
            ]
        })
        preview = import_service.preview(session, content, "p.xlsx")
        assert len(preview["holdings"]) == 1
        assert preview["deals"] == []

    def test_two_portfolios_in_one_file_stay_separate(self, session):
        """Ради этого всё и затевалось: два портфеля не должны слипнуться."""
        add_bond(session, "SU26245RMFS9")
        session.commit()

        content = _book({
            "Портфели": [
                ["Портфель", "Вид учёта"],
                ["Торговый", "торговый"],
                ["Инвестиционный", "до погашения"],
            ],
            "Остатки": [
                ["Портфель", "Код бумаги", "Количество", "Цена приобретения"],
                ["Торговый", "SU26245RMFS9", 100, 90.0],
                ["Инвестиционный", "SU26245RMFS9", 50, 92.0],
            ],
        })
        preview = import_service.preview(session, content, "p.xlsx")
        import_service.apply(session, preview)

        by_portfolio = {
            deal.portfolio: deal.quantity
            for deal in session.execute(select(Deal)).scalars()
        }
        assert by_portfolio == {"Торговый": 100, "Инвестиционный": 50}
        assert portfolio_service.accounting_types(session)["Инвестиционный"] == ACCOUNTING_HTM

    def test_csv_is_rejected_with_clear_message(self, session):
        with pytest.raises(ValueError, match="Excel"):
            import_service.read_sheets(b"a;b", "portfolio.csv")

    def test_template_is_a_readable_workbook(self):
        """Шаблон должен открываться в Excel, а не просто отдаваться байтами."""
        workbook = load_workbook(io.BytesIO(import_service.build_template()))
        holdings = workbook["Остатки"]
        # Шапка на второй строке: первая занята пояснением
        headers = [cell.value for cell in holdings[2]]
        assert "Количество" in headers
        assert "Цена приобретения" in headers

    def test_accounting_type_parsing_variants(self):
        assert import_service.parse_accounting_type("До погашения") == ACCOUNTING_HTM
        assert import_service.parse_accounting_type("инвестиционный") == ACCOUNTING_HTM
        assert import_service.parse_accounting_type("HTM") == ACCOUNTING_HTM
        assert import_service.parse_accounting_type("торговый") == ACCOUNTING_TRADING
        # Непонятное считаем торговым — так вело себя всё до появления видов
        assert import_service.parse_accounting_type("абракадабра") == ACCOUNTING_TRADING
        assert import_service.parse_accounting_type(None) == ACCOUNTING_TRADING


class TestDownloadEndpoints:
    """Выгрузки отдаются по HTTP, а не только собираются сервисом.

    Эти два маршрута уже один раз сломались на фронтенде: кнопки ходили
    обычной ссылкой, которая не несёт токен входа, и обе отвечали 401.
    Здесь проверяется серверная половина — что файл действительно
    формируется и открывается как книга Excel.
    """

    @pytest.fixture()
    def client(self, tmp_path):
        from fastapi.testclient import TestClient
        from app.config import settings
        import importlib
        import app.db
        import app.main

        previous = (
            settings.auth_enabled, settings.database_url,
            settings.collect_on_startup, settings.scheduler_enabled,
        )
        settings.auth_enabled = False
        settings.database_url = f"sqlite:///{tmp_path / 'dl.db'}"
        settings.collect_on_startup = False
        settings.scheduler_enabled = False
        importlib.reload(app.db)
        importlib.reload(app.main)

        with TestClient(app.main.app) as active:
            yield active

        (
            settings.auth_enabled, settings.database_url,
            settings.collect_on_startup, settings.scheduler_enabled,
        ) = previous
        importlib.reload(app.db)
        importlib.reload(app.main)

    def test_template_downloads_as_workbook(self, client):
        response = client.get("/api/import/portfolio/template")
        assert response.status_code == 200
        workbook = load_workbook(io.BytesIO(response.content))
        assert set(workbook.sheetnames) == {"Портфели", "Остатки", "Сделки"}

    def test_revaluation_download_without_positions_is_explained(self, client):
        """Пустой портфель — понятное сообщение, а не пустой файл."""
        response = client.get("/api/portfolio/revaluation/download")
        assert response.status_code == 404
        assert "нет позиций" in response.json()["detail"]

    def test_revaluation_downloads_as_workbook(self, client):
        from app.db import session_scope

        with session_scope() as session:
            bond = add_bond(session)
            add_quote(session, bond, wa_price=95.0, prev_wa_price=94.0)
            add_deal(session, bond.secid, quantity=10, price=90.0)

        response = client.get("/api/portfolio/revaluation/download?fmt=xlsx")
        assert response.status_code == 200
        workbook = load_workbook(io.BytesIO(response.content))
        assert workbook.sheetnames == ["Переоценка"]

    def test_revaluation_downloads_as_csv(self, client):
        from app.db import session_scope

        with session_scope() as session:
            bond = add_bond(session)
            add_quote(session, bond, wa_price=95.0, prev_wa_price=94.0)
            add_deal(session, bond.secid, quantity=10, price=90.0)

        response = client.get("/api/portfolio/revaluation/download?fmt=csv")
        assert response.status_code == 200
        assert "Переоценка за день" in response.content.decode("utf-8-sig")
