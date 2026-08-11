"""Тесты выгрузки по списку бумаг и анализа облигаций."""
from __future__ import annotations

from datetime import date, datetime

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.models import Base, CorpAction, Instrument, Quote
from app.services import bonds as bonds_service
from app.services import export as export_service
from app.services.export import Resolved
from app.services.tabular import to_csv, to_xlsx


@pytest.fixture()
def session():
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False)
    with factory() as active:
        yield active


class TestParseIdentifiers:
    def test_splits_excel_column(self):
        raw = "RU000A105SG2\nSBER\nSU26238RMFS4"
        assert export_service.parse_identifiers(raw) == [
            "RU000A105SG2", "SBER", "SU26238RMFS4",
        ]

    def test_splits_excel_row_and_separators(self):
        raw = "RU000A105SG2\tSBER, GAZP; LKOH"
        assert export_service.parse_identifiers(raw) == [
            "RU000A105SG2", "SBER", "GAZP", "LKOH",
        ]

    def test_removes_duplicates_and_case(self):
        assert export_service.parse_identifiers("sber\nSBER\n sber ") == ["SBER"]

    def test_strips_quotes_from_pasted_cells(self):
        assert export_service.parse_identifiers('"SBER"\n\'GAZP\'') == ["SBER", "GAZP"]

    def test_empty_input(self):
        assert export_service.parse_identifiers("") == []
        assert export_service.parse_identifiers("   \n  ") == []

    def test_caps_length(self):
        raw = "\n".join(f"SEC{index}" for index in range(500))
        assert len(export_service.parse_identifiers(raw)) == export_service.MAX_SECURITIES

    @pytest.mark.parametrize(
        ("value", "expected"),
        [("RU000A105SG2", True), ("SBER", False), ("RU000A105SG", False)],
    )
    def test_isin_detection(self, value, expected):
        assert export_service.looks_like_isin(value) is expected


class TestBuildTable:
    ITEM = Resolved(
        query="RU000A105SG2",
        secid="RU000A105SG2",
        isin="RU000A105SG2",
        name="ГазКЗ-30Д",
        board="TQCB",
        engine="stock",
        market="bonds",
        kind="bond",
    )
    BARS = [
        {"trade_date": date(2026, 8, 3), "wa_price": 85.0, "close": 85.5,
         "volume": 100.0, "turnover": 1000.0, "accrued_interest": 1133.0,
         "currency": "SUR", "num_trades": 10},
        {"trade_date": date(2026, 8, 4), "wa_price": 84.0, "close": 84.5,
         "volume": 200.0, "turnover": 2000.0, "accrued_interest": 1148.0,
         "currency": "SUR", "num_trades": 20},
    ]

    def test_by_date_rows(self):
        rows = export_service.build_rows(
            self.ITEM, self.BARS, ["wa_price", "accrued_interest"], "by_date"
        )
        assert len(rows) == 2
        assert rows[0]["trade_date"] == date(2026, 8, 3)
        assert rows[0]["wa_price"] == 85.0
        assert rows[0]["secid"] == "RU000A105SG2"

    def test_summary_aggregates(self):
        rows = export_service.build_rows(
            self.ITEM, self.BARS, ["wa_price", "volume", "accrued_interest"], "summary"
        )
        assert len(rows) == 1
        row = rows[0]
        assert row["days"] == 2
        # Средневзвешенная цена сворачивается в среднюю, минимум и максимум
        assert row["wa_price_avg"] == pytest.approx(84.5)
        assert row["wa_price_min"] == pytest.approx(84.0)
        assert row["wa_price_max"] == pytest.approx(85.0)
        # Объём суммируется, НКД берётся на конец периода
        assert row["volume"] == pytest.approx(300.0)
        assert row["accrued_interest"] == pytest.approx(1148.0)

    def test_summary_of_empty_history(self):
        assert export_service.build_rows(self.ITEM, [], ["wa_price"], "summary") == []

    def test_accrued_filled_for_every_day_of_period(self):
        """НКД накапливается каждый день, а не только в дни торгов.

        В истории торгов сегодняшнего дня нет — итоги публикуются назавтра.
        Но купон уже накопился, и строка за сегодня должна быть, пусть цены в
        ней и пустые.
        """
        period = [date(2026, 8, 3), date(2026, 8, 4), date(2026, 8, 5)]
        accrual = {
            date(2026, 8, 3): {"accrued_today": 1133.0, "accrued_settlement": 1148.0},
            date(2026, 8, 4): {"accrued_today": 1148.0, "accrued_settlement": 1163.0},
            date(2026, 8, 5): {"accrued_today": 1163.0, "accrued_settlement": 1178.0},
        }
        rows = export_service.build_rows(
            self.ITEM, self.BARS,
            ["wa_price", "accrued_interest", "accrued_today", "accrued_settlement"],
            "by_date", accrual, dates=period,
        )

        assert len(rows) == 3
        # Торгов 5 августа не было: цена пустая, НКД на месте
        last = rows[-1]
        assert last["trade_date"] == date(2026, 8, 5)
        assert last["no_trades"] is True
        assert last["wa_price"] is None
        assert last["accrued_interest"] is None
        assert last["accrued_today"] == 1163.0
        assert last["accrued_settlement"] == 1178.0
        # В дни с торгами рыночные колонки заполнены
        assert rows[0]["no_trades"] is False
        assert rows[0]["wa_price"] == 85.0

    def test_accrued_changes_day_to_day(self):
        """Одно и то же число во всех строках было бы неправдой."""
        period = [date(2026, 8, 3), date(2026, 8, 4)]
        accrual = {
            date(2026, 8, 3): {"accrued_today": 1133.0},
            date(2026, 8, 4): {"accrued_today": 1148.0},
        }
        rows = export_service.build_rows(
            self.ITEM, self.BARS, ["accrued_today"], "by_date", accrual, dates=period
        )
        assert [row["accrued_today"] for row in rows] == [1133.0, 1148.0]

    def test_summary_takes_accrued_on_last_day(self):
        """В своде расчётный НКД берётся на конец периода."""
        accrual = {
            date(2026, 8, 3): {"accrued_today": 1133.0},
            date(2026, 8, 4): {"accrued_today": 1148.0},
        }
        rows = export_service.build_rows(
            self.ITEM, self.BARS, ["accrued_today"], "summary", accrual
        )
        assert rows[0]["accrued_today"] == 1148.0

    def test_accrued_without_data_is_empty(self):
        """Если купоны неизвестны, честнее пусто, чем ноль."""
        rows = export_service.build_rows(
            self.ITEM, self.BARS, ["accrued_today"], "by_date", {},
            dates=[date(2026, 8, 3)],
        )
        assert rows[0]["accrued_today"] is None

    def test_accrued_column_has_no_aggregation_suffix(self):
        """В своде подпись «на конец» была бы неправдой."""
        columns = export_service.build_columns(
            ["accrued_interest", "accrued_today"], "summary"
        )
        titles = {column["code"]: column["title"] for column in columns}
        assert titles["accrued_today"] == "НКД на дату"
        assert titles["accrued_interest"].startswith("НКД на дату торгов,")

    def test_settlement_skips_weekend(self):
        """Расчёты по пятничной сделке проходят в понедельник, не в субботу."""
        friday = date(2026, 8, 7)
        assert export_service._next_business_day(friday) == date(2026, 8, 10)
        assert export_service._next_business_day(date(2026, 8, 10)) == date(2026, 8, 11)

    def test_date_range_covers_whole_period(self):
        days = export_service.date_range(date(2026, 8, 3), date(2026, 8, 6))
        assert days == [date(2026, 8, i) for i in (3, 4, 5, 6)]
        # Порядок дат перепутан — период всё равно должен получиться
        assert export_service.date_range(date(2026, 8, 6), date(2026, 8, 3)) == days

    def test_accrued_today_offered_in_catalog(self):
        catalog = export_service.parameter_catalog()
        bonds = next(group for group in catalog if group["group"] == "Облигации")
        codes = {item["code"] for item in bonds["items"]}
        assert {"accrued_interest", "accrued_today"} <= codes
        # Подсказка обязана объяснять, чем колонки отличаются
        today = next(i for i in bonds["items"] if i["code"] == "accrued_today")
        assert today["hint"]

    def test_columns_match_mode(self):
        by_date = export_service.build_columns(["wa_price"], "by_date")
        assert [column["code"] for column in by_date] == [
            "secid", "isin", "name", "trade_date", "wa_price",
        ]
        summary = export_service.build_columns(["wa_price"], "summary")
        codes = [column["code"] for column in summary]
        assert "wa_price_avg" in codes and "wa_price_min" in codes
        assert "trade_date" not in codes

    def test_removed_parameters_are_gone(self):
        """Колонки, от которых отказались, не должны вернуться незаметно."""
        codes = {param.code for param in export_service.PARAMS}
        assert not codes & {
            "legal_close", "yield_close", "coupon_percent", "currency",
            "settle_date", "accrued_basis",
        }

    def test_unknown_parameter_is_ignored(self):
        """Сохранённый отбор может ссылаться на убранный параметр."""
        rows = export_service.build_rows(
            self.ITEM, self.BARS, ["currency", "wa_price"], "summary"
        )
        assert "currency" not in rows[0]
        assert rows[0]["wa_price_avg"] is not None


class TestFiles:
    COLUMNS = [
        {"code": "secid", "title": "Код", "kind": "text"},
        {"code": "trade_date", "title": "Дата", "kind": "date"},
        {"code": "wa_price", "title": "СВЦ", "kind": "number", "digits": 4},
    ]
    ROWS = [
        {"secid": "SBER", "trade_date": date(2026, 8, 3), "wa_price": 283.3960},
        {"secid": "SBER", "trade_date": date(2026, 8, 4), "wa_price": None},
    ]

    def test_xlsx_types_and_layout(self, tmp_path):
        content = to_xlsx(self.COLUMNS, self.ROWS, meta=[("Период", "август")])
        path = tmp_path / "out.xlsx"
        path.write_bytes(content)

        sheet = load_workbook(path).active
        # Одна строка метаданных, пустая строка, затем шапка
        assert sheet.cell(row=1, column=1).value == "Период"
        assert sheet.cell(row=3, column=1).value == "Код"
        assert sheet.cell(row=4, column=1).value == "SBER"
        # Числа остаются числами, даты датами — Excel сможет считать по ним
        assert isinstance(sheet.cell(row=4, column=3).value, float)
        assert sheet.cell(row=4, column=3).number_format == "# ##0.0000"
        assert isinstance(sheet.cell(row=4, column=2).value, datetime)
        assert sheet.freeze_panes == "A4"
        # Пустое значение остаётся пустым, а не превращается в ноль
        assert sheet.cell(row=5, column=3).value is None

    def test_csv_is_excel_friendly(self):
        content = to_csv(self.COLUMNS, self.ROWS)
        text = content.decode("utf-8")
        # BOM, разделитель «;» и десятичная запятая
        assert text.startswith("﻿")
        assert "Код;Дата;СВЦ" in text
        assert "283,3960" in text
        assert "03.08.2026" in text

    def test_csv_keeps_empty_cells_empty(self):
        text = to_csv(self.COLUMNS, self.ROWS).decode("utf-8")
        assert text.strip().splitlines()[-1].endswith(";")


class TestCouponProfile:
    """Тип купона берём из справочника MOEX — он есть у всех выпусков."""

    @pytest.mark.parametrize(
        ("bond_type", "expected"),
        [
            ("Фикс с известным купоном", bonds_service.COUPON_FIXED),
            ("Фикс с неизвестным купоном", bonds_service.COUPON_FIXED),
            ("Флоатер", bonds_service.COUPON_FLOAT),
            ("Структурная облигация", bonds_service.COUPON_STRUCTURED),
            ("Линкер/облигации с индексируемым", bonds_service.COUPON_LINKER),
            ("Дисконтная облигация", bonds_service.COUPON_DISCOUNT),
            ("Амортизируемые облигации", bonds_service.COUPON_FIXED),
            (None, bonds_service.COUPON_UNKNOWN),
        ],
    )
    def test_type_from_moex(self, bond_type, expected):
        assert bonds_service._coupon_type_from_moex(bond_type) == expected

    def test_amortization_from_bond_type(self):
        instrument = Instrument(
            secid="X", board="TQCB", engine="stock", market="bonds", kind="bond",
            bond_type="Амортизируемые облигации",
        )
        profile = bonds_service._coupon_profile([], instrument)
        assert profile["has_amortization"] is True

    def test_amortization_from_schedule(self):
        instrument = Instrument(
            secid="X", board="TQCB", engine="stock", market="bonds", kind="bond",
            bond_type="Фикс с известным купоном",
        )
        actions = [
            CorpAction(isin="X", action_type="amortization", action_date=date(2030, 1, 1)),
            CorpAction(isin="X", action_type="amortization", action_date=date(2031, 1, 1)),
        ]
        assert bonds_service._coupon_profile(actions, instrument)["has_amortization"] is True

    def test_single_redemption_is_not_amortization(self):
        instrument = Instrument(
            secid="X", board="TQCB", engine="stock", market="bonds", kind="bond",
            bond_type="Фикс с известным купоном",
        )
        actions = [
            CorpAction(isin="X", action_type="amortization", action_date=date(2030, 1, 1)),
        ]
        assert bonds_service._coupon_profile(actions, instrument)["has_amortization"] is False


class TestRiskScore:
    def test_quality_issue_scores_low(self):
        row = {"spread_to_curve_bp": 60, "list_level": 1, "liquidity_score": 85}
        result = bonds_service._risk_score(row)
        assert result["risk_band"] == "низкий"

    def test_distressed_issue_scores_high(self):
        row = {"spread_to_curve_bp": 2000, "list_level": 3, "liquidity_score": 10}
        result = bonds_service._risk_score(row)
        assert result["risk_band"] == "высокий"
        assert any("премия" in reason for reason in result["risk_reasons"])

    def test_score_is_bounded(self):
        row = {"spread_to_curve_bp": 99999, "list_level": 3, "liquidity_score": 0}
        assert 0 <= bonds_service._risk_score(row)["risk_score"] <= 100

    def test_missing_trading_data_penalised(self):
        result = bonds_service._risk_score({"spread_to_curve_bp": None, "list_level": 1})
        assert any("нет данных" in reason for reason in result["risk_reasons"])


class TestAnalyse:
    def _bond(self, session, secid, **kwargs):
        instrument = Instrument(
            secid=secid, board="TQCB", engine="stock", market="bonds", kind="bond",
            short_name=secid, isin=f"RU{secid}", face_value=1000.0,
            bond_type=kwargs.pop("bond_type", "Фикс с известным купоном"),
            maturity_date=kwargs.pop("maturity_date", date(2030, 1, 1)),
            **kwargs,
        )
        session.add(instrument)
        session.flush()
        return instrument

    def test_filters_by_coupon_type(self, session):
        fixed = self._bond(session, "FIX")
        floater = self._bond(session, "FLT", bond_type="Флоатер")
        for instrument in (fixed, floater):
            session.add(Quote(
                instrument_id=instrument.id, ts=datetime(2026, 8, 10, 12, 0),
                last=100.0, turnover=1e7, yield_pct=15.0, duration_days=365,
            ))
        session.flush()

        result = bonds_service.analyse(session, coupon_types=["float"])
        assert [row["secid"] for row in result["items"]] == ["FLT"]

    def test_filters_by_maturity_window(self, session):
        near = self._bond(session, "NEAR", maturity_date=date(2027, 1, 1))
        far = self._bond(session, "FAR", maturity_date=date(2035, 1, 1))
        for instrument in (near, far):
            session.add(Quote(
                instrument_id=instrument.id, ts=datetime(2026, 8, 10, 12, 0),
                last=100.0, turnover=1e7,
            ))
        session.flush()

        result = bonds_service.analyse(session, maturity_to=date(2028, 1, 1))
        assert [row["secid"] for row in result["items"]] == ["NEAR"]

    def test_computes_current_yield(self, session):
        instrument = self._bond(session, "CUR", coupon_percent=10.0)
        session.add(Quote(
            instrument_id=instrument.id, ts=datetime(2026, 8, 10, 12, 0),
            last=50.0, turnover=1e7,
        ))
        session.flush()

        row = bonds_service.analyse(session)["items"][0]
        # Купон 10% к цене 50% от номинала даёт текущую доходность 20%
        assert row["current_yield_pct"] == pytest.approx(20.0)

    def test_export_rows_use_readable_flags(self, session):
        instrument = self._bond(session, "AMR", bond_type="Амортизируемые облигации")
        session.add(Quote(
            instrument_id=instrument.id, ts=datetime(2026, 8, 10, 12, 0),
            last=100.0, turnover=1e7,
        ))
        session.flush()

        items = bonds_service.analyse(session)["items"]
        prepared = bonds_service.rows_for_export(items)
        assert prepared[0]["has_amortization"] == "да"
