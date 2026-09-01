"""Платёжный календарь: разноска лицевых счетов, матрица и импорт выгрузки.

Проверяем то, где ошибка тише всего: разноска по номеру счёта (сумма уходит
не в ту статью и растворяется), знак остатка в оборотной ведомости (актив и
пассив нельзя складывать), повторная загрузка того же дня (обороты удвоятся)
и накопительное сальдо (по нему решают, хватит ли денег).
"""
from __future__ import annotations

import io
from datetime import date, timedelta

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.models import Base, LedgerRow
from app.services import calendar_matrix as matrix_service
from app.services import ledger_import as ledger_service

DAY = date(2026, 3, 17)


@pytest.fixture()
def session():
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False)
    with factory() as active:
        yield active


def build_export(
    rows,
    *,
    header_note: str | None = "Оборотная ведомость по состоянию на 17.03.2026",
    split_balance: bool = False,
) -> bytes:
    """Собрать оборотную ведомость, похожую на банковскую выгрузку."""
    workbook = Workbook()
    sheet = workbook.active
    line = 1
    if header_note:
        sheet.cell(row=line, column=1, value=header_note)
        line += 2

    if split_balance:
        headers = [
            "Лицевой счёт", "Наименование", "Валюта",
            "Входящий остаток по дебету", "Входящий остаток по кредиту",
            "Оборот по дебету", "Оборот по кредиту",
            "Исходящий остаток по дебету", "Исходящий остаток по кредиту",
        ]
    else:
        headers = [
            "Лицевой счёт", "Наименование", "Валюта",
            "Входящий остаток", "Оборот по дебету", "Оборот по кредиту",
            "Исходящий остаток",
        ]
    for index, title in enumerate(headers, start=1):
        sheet.cell(row=line, column=index, value=title)

    for offset, row in enumerate(rows, start=line + 1):
        for index, value in enumerate(row, start=1):
            sheet.cell(row=offset, column=index, value=value)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


SAMPLE = [
    # счёт, наименование, валюта, вх. остаток, дебет, кредит, исх. остаток
    ("30102810000000000123", "Корсчёт в Банке России", "RUB", 5_000_000, 0, 0, 5_000_000),
    ("60305810400000000001", "Расчёты с работниками", "RUB", 0, 900_000, 0, 0),
    ("60301810800000000100", "Единый налоговый платёж", "RUB", 0, 300_000, 0, 0),
    ("60306810100000000005", "Расчёты с поставщиками", "RUB", 0, 120_000, 0, 0),
    ("32002810700204900187", "МБК размещённый", "RUB", 0, 2_000_000, 0, 2_000_000),
    ("31902810600000000000", "Депозит в Банке России", "RUB", 0, 1_500_000, 0, 1_500_000),
    ("45815810100000000003", "Просроченная задолженность ФЛ", "RUB", 0, 0, 750_000, 0),
    ("40702810500000000777", "Расчётный счёт клиента", "RUB", 0, 400_000, 1_100_000, -700_000),
    ("42102810900000000010", "Депозит юрлица", "RUB", 0, 0, 3_000_000, -3_000_000),
    ("Итого по разделу", "", "", 0, 3_720_000, 4_850_000, 0),
]


# ----------------------------------------------------------------------
# Раскладка эталонного файла
# ----------------------------------------------------------------------
#: Столбец A рабочего файла казначейства, строки 2–55, слово в слово.
#:
#: Этот список — договор с казначейством, а не украшение теста: календарь
#: сверяют глазами с прежним файлом, и переставленная, переименованная или
#: потерянная строка ломает сверку молча. Пустые строки значимы так же, как
#: заполненные: по ним глаз находит границу блока.
TEMPLATE_COLUMN_A: tuple[str, ...] = (
    "На начало дня Остаток на кор. счете 30102",
    "Поступления денежных средств",
    "1. Кредиты ФЛ",
    "1.1. Кредиты ЮЛ",
    "1.2. Кредиты ФЛ Цессия",
    "2.1. МБК_тело",
    "2.2. МБК_ проценты",
    "3.1. СФУК_тело",
    "3.2. СФУК_проценты",
    "4. Депозит ЮЛ_тело",
    "5. Депозит ЦБ",
    "5.2. Депозит в Экспо",
    "5.1. Депозит ЦБ_тело",
    "5.2. Депозит ЦБ_проценты",
    "6. Движения по расчетным счетам клиентов (Поступления)",
    "7. Прочее",
    "ИТОГО поступлений",
    "",
    "Планируемые Платежи",
    "1. Кредиты ФЛ Цессия",
    "1.1. Кредиты ЮЛ",
    "2.1. МБК_тело",
    "2.2. МБК_проценты",
    "3. СФУК_тело",
    "4.1. Депозит ЮЛ _тело",
    "4.2. Депозиты ЮЛ_проценты",
    "5.1. Депозит ЦБ",
    "5.2. Депозит в Экспо",
    "6. Движения по расчетным счетам клиентов (Списания)",
    "7. Зарплата",
    "8. Платежи общехозяйственные (поставщикам услуг через WSS)",
    "9. Налоги",
    "10. Налоги зп",
    "11. Прочее",
    "ИТОГО платежей",
    "",
    "САЛЬДО ДНЯ",
    "в т.ч. ФОР",
    "Накопительное сальдо",
    "",
    "Остатки на клиентских счетах",
    "Счет 407, Счет 408",
    "Счет 420, Счет 421",
    "",
    "Н2 (триггер не менее 17,5)",
    "",
    "Н3 (триггер не менее 57,50)",
    "",
    "Для Retail и цессия (стар)",
    "Для 1-го транша (не удалять)",
    "Для 2-го транша (не удалять)",
    "Остаток со сроком 12+ мес.",
    "Капитал",
    "Н4",
)


class TestTemplateLayout:
    def test_rows_repeat_the_template_column_word_for_word(self):
        assert tuple(row.title for row in matrix_service.ROWS) == TEMPLATE_COLUMN_A

    def test_captions_and_spacers_are_rows_not_decoration(self):
        """В файле это настоящие строки — терминал не должен их выдумывать."""
        kinds = {row.title: row.kind for row in matrix_service.ROWS}
        assert kinds["Поступления денежных средств"] == matrix_service.KIND_CAPTION
        assert kinds["Планируемые Платежи"] == matrix_service.KIND_CAPTION
        assert kinds["Остатки на клиентских счетах"] == matrix_service.KIND_CAPTION
        assert kinds[""] == matrix_service.KIND_SPACER

    def test_no_invented_section_headings(self):
        """Над «САЛЬДО ДНЯ» и над блоком нормативов в файле заголовка нет."""
        titles = [row.title for row in matrix_service.ROWS]
        index = titles.index("САЛЬДО ДНЯ")
        assert titles[index - 1] == ""
        assert titles.index("Н2 (триггер не менее 17,5)") - 1 == titles.index(
            "", titles.index("Счет 420, Счет 421")
        )

    def test_row_codes_are_unique(self):
        codes = [row.code for row in matrix_service.ROWS]
        assert len(codes) == len(set(codes))

    def test_only_articles_carry_values(self, session):
        result = matrix_service.matrix(session)
        for row in result["rows"]:
            if row["kind"] == matrix_service.KIND_ARTICLE:
                assert len(row["values"]) == len(result["days"]), row["title"]
            else:
                assert row["values"] == [], row["title"]

    def test_accent_matches_the_template_fill(self):
        """Голубая заливка файла: строка начала дня, поступления и оба ИТОГО."""
        accented = {row.title for row in matrix_service.ROWS if row.accent}
        assert "На начало дня Остаток на кор. счете 30102" in accented
        assert "ИТОГО поступлений" in accented
        assert "ИТОГО платежей" in accented
        # Строки 31–35 файла залиты не были — казначейство ведёт их руками
        assert "7. Зарплата" not in accented
        assert "9. Налоги" not in accented

    def test_running_total_is_the_green_row(self):
        running = [row.title for row in matrix_service.ROWS if row.running]
        assert running == ["Накопительное сальдо"]


# ----------------------------------------------------------------------
# Разноска счетов
# ----------------------------------------------------------------------
class TestClassifier:
    def test_salary_account_beats_general_603_rule(self):
        """Частное правило должно перебивать общее, а не наоборот."""
        rule = matrix_service.classify("60305810400000000001", matrix_service.DEBIT)
        assert rule.row_code == "out_salary"

    def test_other_603_goes_to_overhead(self):
        rule = matrix_service.classify("60306810100000000005", matrix_service.DEBIT)
        assert rule.row_code == "out_overhead"

    def test_single_tax_payment_account(self):
        rule = matrix_service.classify("60301810800000000100", matrix_service.DEBIT)
        assert rule.row_code == "out_taxes"

    def test_same_account_maps_differently_by_side(self):
        """458 по дебету — прочий платёж, по кредиту — возврат кредита."""
        account = "45815810100000000003"
        assert (
            matrix_service.classify(account, matrix_service.DEBIT).row_code == "out_other"
        )
        assert (
            matrix_service.classify(account, matrix_service.CREDIT).row_code
            == "in_loans_retail"
        )

    def test_interest_accounts_go_to_deposit_interest(self):
        for account in ("47426810900000000012", "47422810300000000004"):
            rule = matrix_service.classify(account, matrix_service.DEBIT)
            assert rule.row_code == "out_deposit_corp_interest", account

    def test_corr_account_is_not_an_article(self):
        """30102 — остаток на начало дня, а не оборот статьи."""
        assert matrix_service.classify("30102810000000000123", matrix_service.DEBIT) is None

    def test_spaces_and_dots_in_account_number(self):
        rule = matrix_service.classify("407 02.810 5000 0000 0777", matrix_service.CREDIT)
        assert rule.row_code == "in_client_accounts"

    def test_unknown_account_has_no_rule(self):
        assert matrix_service.classify("99999810000000000001", matrix_service.DEBIT) is None

    def test_every_rule_points_at_a_real_row(self):
        for rule in matrix_service.RULES:
            assert rule.row_code in matrix_service.ROW_BY_CODE, rule.prefix

    def test_rules_table_is_ordered_by_specificity(self):
        lengths = [len(item["prefix"]) for item in matrix_service.rules_table()]
        assert lengths == sorted(lengths, reverse=True)


# ----------------------------------------------------------------------
# Чтение выгрузки
# ----------------------------------------------------------------------
class TestParse:
    def test_reads_accounts_and_turnovers(self):
        parsed = ledger_service.parse(build_export(SAMPLE), "оборотка.xlsx")
        accounts = {row["account"] for row in parsed["rows"]}
        assert "30102810000000000123" in accounts
        salary = next(
            row for row in parsed["rows"] if row["account"] == "60305810400000000001"
        )
        assert salary["debit_turnover"] == 900_000
        assert salary["debit_row"] == "out_salary"

    def test_total_line_without_account_is_skipped(self):
        """«Итого по разделу» не счёт — иначе обороты задвоятся."""
        parsed = ledger_service.parse(build_export(SAMPLE), "оборотка.xlsx")
        assert len(parsed["rows"]) == len(SAMPLE) - 1
        assert parsed["skipped"] == 1

    def test_date_is_read_from_the_report_header(self):
        parsed = ledger_service.parse(build_export(SAMPLE), "оборотка.xlsx")
        assert parsed["load_date"] == DAY
        assert parsed["detected_date"] == DAY

    def test_explicit_date_wins_over_the_header(self):
        """В шапке может стоять дата печати, поэтому человек главнее файла."""
        chosen = date(2026, 4, 1)
        parsed = ledger_service.parse(
            build_export(SAMPLE), "оборотка.xlsx", on_date=chosen
        )
        assert parsed["load_date"] == chosen

    def test_falls_back_to_today_without_a_date(self):
        parsed = ledger_service.parse(
            build_export(SAMPLE, header_note=None), "оборотка.xlsx"
        )
        assert parsed["load_date"] == date.today()
        assert parsed["detected_date"] is None

    def test_split_balance_columns_keep_the_sign(self):
        """Дебетовый и кредитовый остатки нельзя складывать как есть."""
        rows = [
            ("30102810000000000123", "Корсчёт", "RUB", 5_000_000, 0, 0, 0, 5_000_000, 0),
            ("42102810900000000010", "Депозит", "RUB", 0, 1_000_000, 0, 0, 0, 1_000_000),
        ]
        parsed = ledger_service.parse(
            build_export(rows, split_balance=True), "оборотка.xlsx"
        )
        by_account = {row["account"]: row for row in parsed["rows"]}
        assert by_account["30102810000000000123"]["opening_balance"] == 5_000_000
        assert by_account["42102810900000000010"]["opening_balance"] == -1_000_000

    def test_split_balance_headers_do_not_steal_the_turnover_columns(self):
        """«Оборот по дебету» и «Входящий остаток по дебету» — разные столбцы."""
        rows = [("40702810500000000777", "Клиент", "RUB", 0, 10, 400_000, 1_100_000, 0, 20)]
        parsed = ledger_service.parse(
            build_export(rows, split_balance=True), "оборотка.xlsx"
        )
        row = parsed["rows"][0]
        assert row["debit_turnover"] == 400_000
        assert row["credit_turnover"] == 1_100_000

    def test_duplicate_account_rows_are_summed(self):
        rows = [
            ("40702810500000000777", "Клиент", "RUB", 0, 100, 0, 0),
            ("40702810500000000777", "Клиент", "RUB", 0, 250, 0, 0),
        ]
        parsed = ledger_service.parse(build_export(rows), "оборотка.xlsx")
        assert len(parsed["rows"]) == 1
        assert parsed["rows"][0]["debit_turnover"] == 350

    def test_file_without_account_column_is_rejected(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Дата", "Сумма"])
        sheet.append([DAY, 100])
        buffer = io.BytesIO()
        workbook.save(buffer)
        with pytest.raises(ValueError, match="лицевого счёта"):
            ledger_service.parse(buffer.getvalue(), "не то.xlsx")

    def test_file_without_turnovers_is_rejected(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Лицевой счёт", "Наименование"])
        sheet.append(["40702810500000000777", "Клиент"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        with pytest.raises(ValueError, match="обороты"):
            ledger_service.parse(buffer.getvalue(), "не то.xlsx")


# ----------------------------------------------------------------------
# Предпросмотр и запись
# ----------------------------------------------------------------------
def first_day(result) -> dict:
    """Суммы первого столбца по кодам статей: заголовки и пустые строки мимо."""
    return {
        row["code"]: row["values"][0]
        for row in result["rows"]
        if row["kind"] == matrix_service.KIND_ARTICLE
    }


def load(session, rows=SAMPLE, *, on_date=DAY):
    content = build_export(rows)
    preview = ledger_service.preview(session, content, "оборотка.xlsx", on_date=on_date)
    ledger_service.apply(session, preview, source_file="оборотка.xlsx")
    return preview


class TestPreviewAndApply:
    def test_preview_groups_amounts_by_article(self, session):
        preview = ledger_service.preview(
            session, build_export(SAMPLE), "оборотка.xlsx", on_date=DAY
        )
        by_code = {item["row_code"]: item["amount"] for item in preview["articles"]}
        assert by_code["out_salary"] == 900_000
        assert by_code["out_taxes"] == 300_000
        assert by_code["in_deposit_corp"] == 3_000_000

    def test_corr_account_turnover_is_not_reported_as_lost(self, session):
        """Оборот 30102 — зеркало всех проводок разом, а не потерянная сумма."""
        rows = [
            ("30102810000000000123", "Корсчёт", "RUB", 5_000_000, 900_000, 750_000, 4_850_000),
        ]
        preview = ledger_service.preview(
            session, build_export(rows), "оборотка.xlsx", on_date=DAY
        )
        assert preview["unmapped_count"] == 0
        assert not any("не попали ни в одну статью" in t for t in preview["warnings"])

    def test_reserve_fund_turnover_is_not_reported_as_lost(self, session):
        rows = [("30202810100000000001", "ФОР", "RUB", 0, 100_000, 0, 3_000_000)]
        preview = ledger_service.preview(
            session, build_export(rows), "оборотка.xlsx", on_date=DAY
        )
        assert preview["unmapped_count"] == 0

    def test_preview_warns_about_accounts_without_an_article(self, session):
        rows = [("99999810000000000001", "Неизвестный", "RUB", 0, 12_345, 0, 0)]
        preview = ledger_service.preview(
            session, build_export(rows), "оборотка.xlsx", on_date=DAY
        )
        assert preview["unmapped_count"] == 1
        assert any("не попали ни в одну статью" in text for text in preview["warnings"])

    def test_apply_writes_one_row_per_account(self, session):
        load(session)
        assert session.query(LedgerRow).count() == len(SAMPLE) - 1

    def test_repeated_load_replaces_the_day(self, session):
        """Выгрузка — полный срез дня: дозапись удвоила бы обороты."""
        load(session)
        load(session)
        assert session.query(LedgerRow).count() == len(SAMPLE) - 1
        result = matrix_service.matrix(session)
        salary = next(row for row in result["rows"] if row["code"] == "out_salary")
        assert salary["values"][0] == 900_000

    def test_reload_warns_that_the_day_is_replaced(self, session):
        load(session)
        preview = ledger_service.preview(
            session, build_export(SAMPLE), "оборотка.xlsx", on_date=DAY
        )
        assert preview["replaces"] is True
        assert any("уже загружена" in text for text in preview["warnings"])

    def test_loading_another_day_keeps_the_first(self, session):
        load(session)
        load(session, on_date=DAY + timedelta(days=1))
        assert matrix_service.loaded_dates(session) == [DAY, DAY + timedelta(days=1)]

    def test_drop_removes_only_that_day(self, session):
        load(session)
        load(session, on_date=DAY + timedelta(days=1))
        removed = ledger_service.drop(session, DAY)
        assert removed == len(SAMPLE) - 1
        assert matrix_service.loaded_dates(session) == [DAY + timedelta(days=1)]


# ----------------------------------------------------------------------
# Матрица
# ----------------------------------------------------------------------
class TestMatrix:
    def test_opening_balance_comes_from_the_corr_account(self, session):
        load(session)
        result = matrix_service.matrix(session)
        opening = next(row for row in result["rows"] if row["code"] == "opening")
        assert opening["values"][0] == 5_000_000

    def test_totals_sum_the_articles(self, session):
        load(session)
        result = matrix_service.matrix(session)
        by_code = first_day(result)
        assert by_code["out_total"] == 900_000 + 300_000 + 120_000 + 2_000_000 \
            + 1_500_000 + 400_000
        assert by_code["in_total"] == 750_000 + 1_100_000 + 3_000_000
        assert by_code["day_net"] == by_code["in_total"] - by_code["out_total"]

    def test_cumulative_starts_from_the_opening_balance(self, session):
        load(session)
        result = matrix_service.matrix(session)
        by_code = first_day(result)
        assert by_code["cumulative"] == 5_000_000 + by_code["day_net"]

    def test_cumulative_carries_across_days(self, session):
        load(session)
        load(session, on_date=DAY + timedelta(days=1))
        result = matrix_service.matrix(session)
        cumulative = next(row for row in result["rows"] if row["code"] == "cumulative")
        net = next(row for row in result["rows"] if row["code"] == "day_net")
        assert cumulative["values"][1] == cumulative["values"][0] + net["values"][1]

    def test_missing_day_stays_empty_not_zero(self, session):
        """Пустой день — «выгрузку не загрузили», а не «оборотов не было»."""
        load(session)
        load(session, on_date=DAY + timedelta(days=2))
        result = matrix_service.matrix(session)
        assert [day["loaded"] for day in result["days"]] == [True, False, True]
        net = next(row for row in result["rows"] if row["code"] == "day_net")
        assert net["values"][1] is None
        assert result["empty_days"] == 1

    def test_only_loaded_hides_gaps(self, session):
        load(session)
        load(session, on_date=DAY + timedelta(days=2))
        result = matrix_service.matrix(session, only_loaded=True)
        assert len(result["days"]) == 2

    def test_client_balances_are_shown_without_sign(self, session):
        load(session)
        result = matrix_service.matrix(session)
        by_code = first_day(result)
        assert by_code["client_407_408"] == 700_000
        assert by_code["client_420_421"] == 3_000_000

    def test_empty_database_still_returns_the_structure(self, session):
        result = matrix_service.matrix(session)
        assert len(result["rows"]) == len(matrix_service.ROWS)
        assert result["loaded_days"] == 0

    def test_reversed_range_is_accepted(self, session):
        load(session)
        result = matrix_service.matrix(
            session, date_from=DAY + timedelta(days=2), date_to=DAY
        )
        assert result["date_from"] == DAY
        assert len(result["days"]) == 3


# ----------------------------------------------------------------------
# Лист «Счета» и выгрузка
# ----------------------------------------------------------------------
class TestLedgerSheet:
    def test_shows_the_article_each_turnover_went_to(self, session):
        load(session)
        sheet = matrix_service.ledger_sheet(session, on_date=DAY)
        by_account = {row["account"]: row for row in sheet["rows"]}
        assert by_account["60305810400000000001"]["debit_title"] == "7. Зарплата"
        assert (
            by_account["45815810100000000003"]["credit_title"] == "1. Кредиты ФЛ"
        )

    def test_defaults_to_the_latest_loaded_day(self, session):
        load(session)
        load(session, on_date=DAY + timedelta(days=1))
        assert matrix_service.ledger_sheet(session)["load_date"] == DAY + timedelta(days=1)

    def test_counts_accounts_the_calendar_will_not_show(self, session):
        load(session, rows=SAMPLE + [("99999810000000000001", "?", "RUB", 0, 5, 0, 0)])
        assert matrix_service.ledger_sheet(session, on_date=DAY)["unmapped"] == 1

    def test_service_accounts_are_marked_not_counted_as_errors(self, session):
        rows = [
            ("30102810000000000123", "Корсчёт", "RUB", 5_000_000, 900_000, 0, 4_100_000),
        ]
        load(session, rows=rows)
        sheet = matrix_service.ledger_sheet(session, on_date=DAY)
        assert sheet["unmapped"] == 0
        assert sheet["rows"][0]["technical"] is True

    def test_workbook_keeps_the_treasury_layout(self, session):
        load(session)
        result = matrix_service.matrix(session)
        ledger = matrix_service.ledger_sheet(session, on_date=DAY)
        book = load_workbook(io.BytesIO(matrix_service.build_workbook(result, ledger)))

        assert book.sheetnames == ["платежный календарь", "Счета"]
        sheet = book["платежный календарь"]
        assert sheet["A1"].value == "Статья/Дата"
        assert sheet["B1"].value.date() == DAY
        assert sheet["A2"].value == "На начало дня Остаток на кор. счете 30102"
        assert sheet["B2"].value == 5_000_000

    def test_workbook_reproduces_the_template_row_for_row(self, session):
        """Строка за строкой, включая пустые: по этому файлу ведут сверку."""
        load(session)
        result = matrix_service.matrix(session)
        book = load_workbook(io.BytesIO(matrix_service.build_workbook(result)))
        sheet = book["платежный календарь"]

        column = tuple(
            sheet.cell(row=index, column=1).value or ""
            for index in range(2, 2 + len(TEMPLATE_COLUMN_A))
        )
        assert column == TEMPLATE_COLUMN_A
        # Ровно как в исходнике: статьи занимают строки 2–55
        assert sheet.cell(row=55, column=1).value == "Н4"

    def test_workbook_keeps_the_template_colours(self, session):
        load(session)
        result = matrix_service.matrix(session)
        book = load_workbook(io.BytesIO(matrix_service.build_workbook(result)))
        sheet = book["платежный календарь"]

        # Голубая заливка на «На начало дня» и «ИТОГО поступлений»
        assert sheet["A2"].fill.fgColor.rgb.endswith("00B0F0")
        assert sheet["A18"].fill.fgColor.rgb.endswith("00B0F0")
        # Зелёная — на накопительном сальдо
        assert sheet["A40"].fill.fgColor.rgb.endswith("92D050")

    def test_workbook_writes_dates_in_russian_order(self, session):
        """mm-dd-yy из исходника в русском календаре читается как другая дата."""
        load(session)
        result = matrix_service.matrix(session)
        book = load_workbook(io.BytesIO(matrix_service.build_workbook(result)))
        assert book["платежный календарь"]["B1"].number_format == "DD.MM.YYYY"

    def test_account_numbers_stay_text_in_the_workbook(self, session):
        """Число съело бы ведущие нули — счёт перестанет искаться поиском."""
        load(session)
        result = matrix_service.matrix(session)
        ledger = matrix_service.ledger_sheet(session, on_date=DAY)
        book = load_workbook(io.BytesIO(matrix_service.build_workbook(result, ledger)))
        assert book["Счета"]["A4"].value == "30102810000000000123"
