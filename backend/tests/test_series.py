"""Ряды графиков обзора рынка: каталог, период, показатели и выгрузка."""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.db import get_session
from app.main import app
from app.models import Bar, Base, FxRate, Instrument, MacroRate
from app.services import series
from app.sources.cbr import CbrSource


@pytest.fixture()
def client(monkeypatch):
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False)

    with factory() as session:
        _seed(session)
        session.commit()

    def override_session():
        session = factory()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_session] = override_session
    monkeypatch.setattr("app.main.settings.collect_on_startup", False)
    monkeypatch.setattr("app.main.settings.scheduler_enabled", False)

    with TestClient(app) as test_client:
        yield test_client

    app.dependency_overrides.clear()


def _seed(session):
    """30 дней ставок, курсов и значений индекса, заканчивая сегодняшним днём."""
    index = Instrument(
        secid="IMOEX", board="SNDX", engine="stock", market="index", kind="index",
        short_name="Индекс МосБиржи",
    )
    session.add(index)
    session.flush()

    today = date.today()
    for offset in range(30):
        day = today - timedelta(days=offset)
        session.add_all([
            MacroRate(code="KEY_RATE", name="Ключевая ставка", value=14.0 + offset * 0.01,
                      rate_date=day),
            MacroRate(code="RUONIA", name="RUONIA", value=13.5 + offset * 0.01,
                      rate_date=day),
            MacroRate(code="RUONIA_VOLUME", name="Объём", value=700.0 + offset,
                      rate_date=day),
            FxRate(source="cbr", code="USD", nominal=1, value=80.0 + offset,
                   rate_date=day),
            FxRate(source="cbr", code="CNY", nominal=10, value=110.0 + offset,
                   rate_date=day),
            Bar(instrument_id=index.id, trade_date=day, close=2700.0 + offset,
                open=2690.0 + offset, turnover=1.5e10),
        ])


class TestCatalog:
    def test_lists_all_charts(self, client):
        payload = client.get("/api/series/catalog").json()
        assert [chart["code"] for chart in payload] == ["rates", "imoex", "usd", "cny"]

    def test_rates_chart_exposes_every_cbr_metric(self, client):
        rates = next(
            chart for chart in client.get("/api/series/catalog").json()
            if chart["code"] == "rates"
        )
        codes = {metric["code"] for metric in rates["metrics"]}
        assert codes == {
            "KEY_RATE", "RUONIA", "RUONIA_MIN", "RUONIA_P25", "RUONIA_P75",
            "RUONIA_MAX", "RUONIA_VOLUME", "RUONIA_DEALS", "RUONIA_PARTICIPANTS",
        }

    def test_only_key_rate_and_ruonia_are_on_by_default(self, client):
        rates = next(
            chart for chart in client.get("/api/series/catalog").json()
            if chart["code"] == "rates"
        )
        defaults = {m["code"] for m in rates["metrics"] if m["default"]}
        assert defaults == {"KEY_RATE", "RUONIA"}

    def test_volume_is_drawn_as_bars_on_the_right_axis(self, client):
        rates = next(
            chart for chart in client.get("/api/series/catalog").json()
            if chart["code"] == "rates"
        )
        volume = next(m for m in rates["metrics"] if m["code"] == "RUONIA_VOLUME")
        assert volume["kind"] == "bar"
        assert volume["axis"] == "right"


class TestSeriesData:
    def test_defaults_to_key_rate_and_ruonia(self, client):
        payload = client.get("/api/series/rates").json()
        assert [serie["code"] for serie in payload["series"]] == ["KEY_RATE", "RUONIA"]

    def test_default_period_is_twelve_months(self, client):
        payload = client.get("/api/series/rates").json()
        span = date.fromisoformat(payload["date_to"]) - date.fromisoformat(
            payload["date_from"]
        )
        assert span.days == series.DEFAULT_PERIOD_DAYS

    def test_selected_metrics_are_returned_in_catalog_order(self, client):
        payload = client.get(
            "/api/series/rates", params={"metric": ["RUONIA_VOLUME", "KEY_RATE"]}
        ).json()
        assert [serie["code"] for serie in payload["series"]] == [
            "KEY_RATE", "RUONIA_VOLUME"
        ]

    def test_period_narrows_the_series(self, client):
        today = date.today()
        payload = client.get(
            "/api/series/rates",
            params={"date_from": (today - timedelta(days=6)).isoformat(),
                    "date_to": today.isoformat()},
        ).json()
        points = payload["series"][0]["points"]
        assert len(points) == 7
        assert points[0]["date"] == (today - timedelta(days=6)).isoformat()

    def test_index_chart_reads_daily_bars(self, client):
        payload = client.get("/api/series/imoex").json()
        assert payload["series"][0]["code"] == "close"
        assert len(payload["series"][0]["points"]) == 30

    def test_fx_is_reported_per_single_unit(self, client):
        """ЦБ котирует юань за 10 единиц — на графике должен быть курс за один."""
        payload = client.get("/api/series/cny").json()
        latest = payload["series"][0]["points"][-1]
        assert latest["value"] == pytest.approx(11.0)

    def test_unknown_chart_gives_404(self, client):
        assert client.get("/api/series/nonexistent").status_code == 404


class TestDownload:
    def test_xlsx_has_a_column_per_metric(self, client):
        response = client.get(
            "/api/series/rates/download",
            params={"metric": ["KEY_RATE", "RUONIA", "RUONIA_VOLUME"]},
        )
        assert response.status_code == 200

        sheet = load_workbook(io.BytesIO(response.content)).active
        header = next(
            row for row in sheet.iter_rows(values_only=True) if row[0] == "Дата"
        )
        assert list(header[:4]) == [
            "Дата", "Ключевая ставка, %", "RUONIA, %", "Объём сделок, млрд ₽"
        ]

    def test_csv_is_offered_too(self, client):
        response = client.get(
            "/api/series/usd/download", params={"fmt": "csv"}
        )
        assert response.status_code == 200
        assert "Курс ЦБ" in response.content.decode("utf-8")

    def test_empty_period_is_not_an_empty_file(self, client):
        response = client.get(
            "/api/series/rates/download",
            params={"date_from": "2000-01-01", "date_to": "2000-02-01"},
        )
        assert response.status_code == 404


class TestTable:
    def test_dates_from_different_metrics_share_a_row(self, client):
        with next(app.dependency_overrides[get_session]()) as session:
            data = series.build(session, "rates", ["KEY_RATE", "RUONIA_VOLUME"])
        columns, rows = series.to_table(data)

        assert [column["code"] for column in columns] == [
            "date", "KEY_RATE", "RUONIA_VOLUME"
        ]
        # Обе метрики заданы на каждый день — значит и колонки заполнены обе
        assert all("KEY_RATE" in row and "RUONIA_VOLUME" in row for row in rows)
        assert rows == sorted(rows, key=lambda row: row["date"])


class TestCbrRuoniaParsing:
    """Разбор страницы динамики RUONIA — источник всех восьми показателей."""

    PAGE = """
      <table class="data">
        <tr><th>Дата</th><th>Ставка</th><th>Объем</th><th>Сделок</th>
            <th>Участников</th><th>Мин</th><th>25%</th><th>75%</th>
            <th>Макс</th><th>Статус</th><th>Публикация</th></tr>
        <tr><td>13.08.2026</td><td class="right">13,98</td><td>677,71</td>
            <td>63</td><td>21</td><td>12,80</td><td>13,75</td><td>14,00</td>
            <td>14,15</td><td>Стандартный</td><td>14.08.2026</td></tr>
      </table>
    """

    @pytest.mark.asyncio
    async def test_every_metric_is_extracted(self, monkeypatch):
        class FakeResponse:
            text = self.PAGE

        async def fake_get(self, path, **params):
            return FakeResponse()

        monkeypatch.setattr(CbrSource, "get", fake_get)

        async with CbrSource() as cbr:
            rows = await cbr.fetch_ruonia_details(date(2026, 8, 1), date(2026, 8, 15))

        values = {row["code"]: row["value"] for row in rows}
        assert values == {
            "RUONIA": 13.98,
            "RUONIA_VOLUME": 677.71,
            "RUONIA_DEALS": 63.0,
            "RUONIA_PARTICIPANTS": 21.0,
            "RUONIA_MIN": 12.80,
            "RUONIA_P25": 13.75,
            "RUONIA_P75": 14.00,
            "RUONIA_MAX": 14.15,
        }
        assert {row["rate_date"] for row in rows} == {date(2026, 8, 13)}

    @pytest.mark.asyncio
    async def test_broken_page_does_not_break_collection(self, monkeypatch):
        async def fake_get(self, path, **params):
            raise RuntimeError("ЦБ недоступен")

        monkeypatch.setattr(CbrSource, "get", fake_get)

        async with CbrSource() as cbr:
            assert await cbr.fetch_ruonia_details() == []
