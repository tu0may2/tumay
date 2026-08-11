"""Отчёт для инвесткомитета: состояние портфеля одной книгой Excel."""
from __future__ import annotations

from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .analytics import market_overview
from .benchmark import compare_portfolio
from .cash import cash_position, payment_calendar
from .limits import check_limits
from .portfolio import portfolio_summary
from .risk import portfolio_cashflow, rate_sensitivity
from .treasury_extras import upcoming_offers

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=13)


def _sheet(workbook: Workbook, title: str, first: bool = False):
    if first:
        sheet = workbook.active
        sheet.title = title[:31]
        return sheet
    return workbook.create_sheet(title[:31])


def _write_table(
    sheet, columns: list[dict[str, Any]], rows: list[dict[str, Any]], start_row: int = 1
) -> int:
    """Записать таблицу и вернуть следующую свободную строку."""
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=start_row, column=index, value=column["title"])
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for offset, row in enumerate(rows, start=1):
        for index, column in enumerate(columns, start=1):
            value = row.get(column["code"])
            cell = sheet.cell(row=start_row + offset, column=index)
            if value is None:
                continue
            if column.get("kind") == "date":
                cell.value = value
                cell.number_format = "DD.MM.YYYY"
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.value = value
                cell.number_format = "# ##0." + "0" * int(column.get("digits", 2)) \
                    if column.get("digits", 2) else "# ##0"
            else:
                cell.value = str(value)

    for index, column in enumerate(columns, start=1):
        longest = len(str(column["title"]))
        for row in rows[:200]:
            value = row.get(column["code"])
            if value is not None:
                longest = max(longest, len(str(value)))
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 12), 44)

    return start_row + len(rows) + 2


def _write_pairs(sheet, title: str, pairs: list[tuple[str, Any]], start_row: int) -> int:
    sheet.cell(row=start_row, column=1, value=title).font = _TITLE_FONT
    for offset, (label, value) in enumerate(pairs, start=1):
        sheet.cell(row=start_row + offset, column=1, value=label)
        cell = sheet.cell(row=start_row + offset, column=2)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.value = value
            cell.number_format = "# ##0.00"
        else:
            cell.value = "—" if value is None else str(value)
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 26
    return start_row + len(pairs) + 2


def build_report(session: Session, *, portfolio: str | None = None) -> bytes:
    """Собрать книгу: сводка, позиции, лимиты, риск, деньги, ориентиры."""
    summary = portfolio_summary(session, portfolio=portfolio)
    cash = cash_position(session, portfolio=portfolio)
    limits = check_limits(session, portfolio=portfolio)
    sensitivity = rate_sensitivity(session, portfolio=portfolio)
    flows = portfolio_cashflow(session, portfolio=portfolio, horizon_days=365)
    calendar = payment_calendar(session, portfolio=portfolio, horizon_days=180)
    benchmark = compare_portfolio(session, portfolio=portfolio, days=90)
    offers = upcoming_offers(session, portfolio=portfolio, horizon_days=180)
    overview = market_overview(session)

    workbook = Workbook()

    # --- Сводка ---
    sheet = _sheet(workbook, "Сводка", first=True)
    row = _write_pairs(sheet, "Отчёт по портфелю", [
        ("Портфель", portfolio or "все"),
        ("Дата отчёта", date.today().strftime("%d.%m.%Y")),
        ("Метод учёта", "ФИФО" if summary["cost_method"] == "fifo" else "по средней"),
    ], 1)

    row = _write_pairs(sheet, "Стоимость и результат", [
        ("Стоимость бумаг, ₽", summary["total_value"]),
        ("Денежная позиция, ₽", cash["total_cash_rub"]),
        ("Размещено, ₽", cash["placed_out_rub"]),
        ("Привлечено, ₽", cash["borrowed_rub"]),
        ("Итого ликвидность, ₽", summary["total_value"] + cash["total_liquidity_rub"]),
        ("Ценовой результат, ₽", summary["price_pnl"]),
        ("Валютный результат, ₽", summary["fx_pnl"]),
        ("Купонный результат, ₽", summary["coupon_result"]),
        ("Реализованный результат, ₽", summary["realized_pnl"]),
        ("Комиссии, ₽", summary["fees"]),
        ("Итого результат, ₽", summary["net_pnl"]),
    ], row)

    row = _write_pairs(sheet, "Риск", [
        ("Дюрация облигаций, лет", summary["weighted_duration_years"]),
        ("Модифицированная дюрация", sensitivity["weighted_modified_duration"]),
        ("Выпуклость", sensitivity["weighted_convexity"]),
        ("Доходность к погашению, %", summary["weighted_yield_pct"]),
        ("Концентрация (HHI)", summary["concentration_hhi"]),
        ("Доля топ-5, %", summary["top5_weight_pct"]),
        ("Лимитов нарушено", limits["breached"]),
    ], row)

    _write_pairs(sheet, "Рыночные ориентиры", [
        ("Ключевая ставка ЦБ, %", (overview.get("key_rate") or {}).get("value")),
        ("RUONIA, %", (overview.get("ruonia") or {}).get("value")),
        ("Средняя ставка размещений, %", cash["weighted_placement_rate"]),
        ("Поступления за год, ₽", flows["total_rub"]),
    ], row)

    # --- Позиции ---
    sheet = _sheet(workbook, "Позиции")
    _write_table(sheet, [
        {"code": "secid", "title": "Код"},
        {"code": "name", "title": "Бумага"},
        {"code": "currency", "title": "Валюта"},
        {"code": "quantity", "title": "Количество", "digits": 0},
        {"code": "avg_price", "title": "Средняя цена", "digits": 4},
        {"code": "last_price", "title": "Текущая цена", "digits": 4},
        {"code": "market_value_rub", "title": "Оценка, ₽"},
        {"code": "price_pnl_rub", "title": "Ценовой P&L, ₽"},
        {"code": "fx_pnl_rub", "title": "Валютный P&L, ₽"},
        {"code": "coupon_result_rub", "title": "Купонный, ₽"},
        {"code": "total_pnl_rub", "title": "Итого, ₽"},
        {"code": "weight_pct", "title": "Доля, %"},
        {"code": "duration_years", "title": "Дюрация, лет"},
        {"code": "yield_pct", "title": "Доходность, %"},
        {"code": "days_to_exit", "title": "Выход, дней", "digits": 1},
    ], summary["positions"])

    # --- Лимиты ---
    sheet = _sheet(workbook, "Лимиты")
    _write_table(sheet, [
        {"code": "kind_title", "title": "Вид лимита"},
        {"code": "subject", "title": "Объект"},
        {"code": "limit_value", "title": "Лимит"},
        {"code": "actual", "title": "Факт"},
        {"code": "utilisation_pct", "title": "Заполнено, %", "digits": 1},
        {"code": "headroom", "title": "Запас"},
        {"code": "status_title", "title": "Статус"},
    ], [
        {**row, "status_title": "нарушен" if row["breached"] else "в норме"}
        for row in limits["items"]
    ])

    # --- Риск ставок ---
    sheet = _sheet(workbook, "Риск ставок")
    _write_table(sheet, [
        {"code": "shift_bp", "title": "Сдвиг, бп", "digits": 0},
        {"code": "impact_rub", "title": "Переоценка, ₽"},
        {"code": "impact_pct", "title": "Переоценка, %", "digits": 3},
        {"code": "convexity_effect_rub", "title": "Вклад выпуклости, ₽"},
    ], sensitivity["scenarios"])

    # --- Деньги ---
    sheet = _sheet(workbook, "Деньги")
    row = _write_table(sheet, [
        {"code": "name", "title": "Счёт"},
        {"code": "bank", "title": "Банк"},
        {"code": "currency", "title": "Валюта"},
        {"code": "balance", "title": "Остаток"},
        {"code": "balance_rub", "title": "Остаток, ₽"},
    ], cash["accounts"])

    _write_table(sheet, [
        {"code": "kind_title", "title": "Вид размещения"},
        {"code": "counterparty", "title": "Контрагент"},
        {"code": "amount", "title": "Сумма"},
        {"code": "currency", "title": "Валюта"},
        {"code": "rate", "title": "Ставка, %"},
        {"code": "start_date", "title": "Начало", "kind": "date"},
        {"code": "end_date", "title": "Окончание", "kind": "date"},
        {"code": "accrued_interest", "title": "Начислено"},
        {"code": "total_at_maturity", "title": "К возврату"},
    ], cash["placements"], start_row=row)

    # --- Календарь ---
    sheet = _sheet(workbook, "Платёжный календарь")
    _write_table(sheet, [
        {"code": "flow_date", "title": "Дата", "kind": "date"},
        {"code": "kind_title", "title": "Тип"},
        {"code": "comment", "title": "Основание"},
        {"code": "amount", "title": "Сумма, ₽"},
        {"code": "balance_after", "title": "Остаток после, ₽"},
    ], calendar["events"])

    # --- Оферты ---
    if offers:
        sheet = _sheet(workbook, "Оферты")
        _write_table(sheet, [
            {"code": "secid", "title": "Код"},
            {"code": "name", "title": "Выпуск"},
            {"code": "offer_date", "title": "Дата оферты", "kind": "date"},
            {"code": "days_left", "title": "Осталось дней", "digits": 0},
            {"code": "quantity", "title": "Количество", "digits": 0},
            {"code": "market_value_rub", "title": "Оценка, ₽"},
            {"code": "source", "title": "Источник"},
        ], offers)

    # --- Сравнение с рынком ---
    sheet = _sheet(workbook, "Сравнение с рынком")
    _write_table(sheet, [
        {"code": "title", "title": "Ориентир"},
        {"code": "return_pct", "title": "Доходность за период, %"},
        {"code": "excess_pct", "title": "Разница, %"},
        {"code": "yield_pct", "title": "Доходность к погашению, %"},
        {"code": "duration_years", "title": "Дюрация, лет"},
    ], [
        {
            "title": "Ваш портфель",
            "return_pct": benchmark["portfolio_return_pct"],
            "excess_pct": None,
            "yield_pct": benchmark["portfolio_yield_pct"],
            "duration_years": benchmark["portfolio_duration_years"],
        },
        *[row for row in benchmark["benchmarks"] if row.get("available")],
    ])

    from io import BytesIO

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
