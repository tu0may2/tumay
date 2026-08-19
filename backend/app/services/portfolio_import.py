"""Импорт портфеля из Excel: портфели, остатки и сделки одной книгой.

Ключевое решение — остатки не хранятся отдельной сущностью, а превращаются во
входящие сделки покупки. Иначе в терминале появилось бы два независимых
источника позиции, и любой расчёт (ФИФО, купонный доход, P&L, лимиты) пришлось
бы учить складывать их между собой — с неизбежным двойным счётом там, где про
это забыли. Одна сделка «входящий остаток» на дату приобретения даёт ту же
позицию и ту же себестоимость, но проходит по всем существующим расчётам без
единой правки.

Книга состоит из трёх листов, каждый необязателен:

* **Портфели** — название и вид учёта (торговый / до погашения);
* **Остатки** — что лежит в портфеле: бумага, количество, цена приобретения;
* **Сделки** — журнал операций, если он ведётся.

Лист опознаётся по названию, а колонки внутри — по заголовкам, поэтому точный
порядок колонок и лишние столбцы значения не имеют.
"""
from __future__ import annotations

import io
from datetime import date
from typing import Any, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from ..models import ACCOUNTING_HTM, ACCOUNTING_TRADING, Deal, Instrument, Portfolio
from .importer import (
    MAX_ROWS,
    detect_columns,
    parse_date,
    parse_number,
    parse_side,
)

#: Как называются листы книги и по каким словам мы их узнаём
SHEET_PORTFOLIOS = "Портфели"
SHEET_HOLDINGS = "Остатки"
SHEET_DEALS = "Сделки"

_SHEET_HINTS: dict[str, tuple[str, ...]] = {
    SHEET_PORTFOLIOS: ("портфел", "счета", "счёта"),
    SHEET_HOLDINGS: ("остат", "позиц", "состав", "портфель на"),
    SHEET_DEALS: ("сделк", "операц", "журнал", "deals"),
}

#: Как в файле могут называть вид учёта. Пишем по-русски, потому что заполнять
#: книгу будет человек, а не программа
_ACCOUNTING_HINTS: tuple[tuple[str, str], ...] = (
    ("до погашения", ACCOUNTING_HTM),
    ("до пога", ACCOUNTING_HTM),
    ("htm", ACCOUNTING_HTM),
    ("инвестицион", ACCOUNTING_HTM),
    ("удержива", ACCOUNTING_HTM),
    ("торгов", ACCOUNTING_TRADING),
    ("trading", ACCOUNTING_TRADING),
    ("спекул", ACCOUNTING_TRADING),
)

ACCOUNTING_TITLES = {
    ACCOUNTING_TRADING: "торговый",
    ACCOUNTING_HTM: "до погашения",
}

#: Колонки листа «Остатки». Подсказки — куски заголовков в нижнем регистре
_HOLDING_HINTS: dict[str, tuple[str, ...]] = {
    "portfolio": ("портфель", "счет", "счёт", "portfolio"),
    "secid": ("код", "тикер", "secid", "бумага", "инструмент", "ticker"),
    "isin": ("isin", "исин"),
    "quantity": ("количество", "кол-во", "штук", "остаток", "шт", "qty"),
    "price": ("цена приобрет", "цена покупк", "цена", "себестоим", "price"),
    "accrued_interest": ("нкд", "накопленн", "accrued"),
    "trade_date": ("дата приобрет", "дата покупк", "дата", "date"),
    "comment": ("коммент", "примеч", "note"),
}


def _norm(value: Any) -> str:
    return str(value or "").strip().lower()


def parse_accounting_type(value: Any) -> str:
    """Вид учёта из текста ячейки. Непонятное считаем торговым.

    Умолчание именно торговое: оно совпадает с тем, как терминал вёл себя до
    появления видов учёта, и не приводит к тому, что бумага вдруг перестала
    переоцениваться из-за опечатки в книге.
    """
    lowered = _norm(value)
    for hint, code in _ACCOUNTING_HINTS:
        if hint in lowered:
            return code
    return ACCOUNTING_TRADING


# ----------------------------------------------------------------------
# Шаблон книги
# ----------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1F4B33")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_NOTE_FONT = Font(italic=True, color="666666")

_TEMPLATE_SHEETS: tuple[tuple[str, tuple[str, ...], tuple[tuple[Any, ...], ...], str], ...] = (
    (
        SHEET_PORTFOLIOS,
        ("Портфель", "Вид учёта", "Комментарий"),
        (
            ("Торговый", "торговый", "Бумаги для перепродажи, переоценка по рынку"),
            ("Инвестиционный", "до погашения", "Держим до погашения, амортизированная стоимость"),
        ),
        "Вид учёта: «торговый» или «до погашения». Влияет на переоценку.",
    ),
    (
        SHEET_HOLDINGS,
        (
            "Портфель",
            "Код бумаги",
            "ISIN",
            "Количество",
            "Цена приобретения",
            "НКД при покупке",
            "Дата приобретения",
            "Комментарий",
        ),
        (
            ("Торговый", "SU26245RMFS9", "RU000A108EG6", 1000, 83.33, 12.45, date(2026, 3, 14), ""),
            ("Инвестиционный", "SBER", "RU0009029540", 500, 271.40, "", date(2026, 1, 20), ""),
        ),
        "Цена облигаций — в процентах от номинала, акций — в рублях. "
        "Достаточно кода ИЛИ ISIN. Остаток превращается в сделку покупки на указанную дату.",
    ),
    (
        SHEET_DEALS,
        (
            "Портфель",
            "Дата",
            "Код бумаги",
            "ISIN",
            "Направление",
            "Количество",
            "Цена",
            "НКД",
            "Комиссия",
            "Контрагент",
        ),
        (
            ("Торговый", date(2026, 4, 2), "SU26245RMFS9", "", "покупка", 500, 83.10, 11.80, 120.5, ""),
            ("Торговый", date(2026, 5, 15), "SBER", "", "продажа", 100, 289.00, "", 45.0, ""),
        ),
        "Заполняйте, если ведёте журнал операций. Лист можно оставить пустым.",
    ),
)


def build_template() -> bytes:
    """Собрать пустую книгу-образец со всеми тремя листами.

    Отдаём файл, а не описание формата в документации: заполнить готовые
    столбцы человек может сразу, а сверять свою таблицу с текстом инструкции
    будет долго и с ошибками.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    for title, headers, examples, note in _TEMPLATE_SHEETS:
        sheet = workbook.create_sheet(title)
        sheet.cell(row=1, column=1, value=note).font = _NOTE_FONT
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 2)
        )

        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=index, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_index, example in enumerate(examples, start=3):
            for col_index, value in enumerate(example, start=1):
                cell = sheet.cell(row=row_index, column=col_index)
                if value == "":
                    continue
                cell.value = value
                if isinstance(value, date):
                    cell.number_format = "DD.MM.YYYY"
                # Примеры выделяем курсивом: их надо заменить, а не дополнить
                cell.font = _NOTE_FONT

        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = max(
                len(header) + 4, 14
            )
        sheet.freeze_panes = sheet.cell(row=3, column=1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ----------------------------------------------------------------------
# Чтение книги
# ----------------------------------------------------------------------
def _classify_by_content(rows: Sequence[Sequence[Any]]) -> str | None:
    """Определить назначение листа по его заголовкам.

    Содержимое надёжнее названия: лист «Мой портфель» со списком бумаг — это
    состав, а не справочник портфелей, хотя по названию похож именно на него.
    Поэтому справочником считаем только лист, где есть вид учёта и при этом
    нет количества: перечень бумаг вида учёта не содержит.
    """
    headers = " ".join(_norm(cell) for row in rows[:10] for cell in row)
    has_quantity = any(
        word in headers for word in ("количество", "кол-во", "штук", "остаток")
    )
    has_side = any(
        word in headers for word in ("направлен", "операц", "сторона", "b/s", "тип сделки")
    )
    has_accounting = any(
        word in headers for word in ("вид учет", "вид учёт", "тип учет", "тип учёт")
    )

    if has_accounting and not has_quantity:
        return SHEET_PORTFOLIOS
    if has_side:
        return SHEET_DEALS
    if has_quantity:
        return SHEET_HOLDINGS
    return None


def read_sheets(content: bytes, filename: str) -> dict[str, list[list[Any]]]:
    """Прочитать книгу и разложить листы по назначению.

    Порядок опознания: точное совпадение названия с нашим (так приходит
    заполненный шаблон), затем содержимое заголовков, и лишь потом похожесть
    названия. Название проверяется последним намеренно — слово «портфель»
    встречается и в заголовке листа с составом.
    """
    name = (filename or "").lower()
    if not name.endswith((".xlsx", ".xlsm", ".xltx")):
        raise ValueError("Портфель импортируется из книги Excel (.xlsx)")

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    canonical = {_norm(title): title for title in _SHEET_HINTS}
    found: dict[str, list[list[Any]]] = {}
    leftovers: list[tuple[str, list[list[Any]]]] = []

    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not any(any(str(cell or "").strip() for cell in row) for row in rows):
            continue

        kind = canonical.get(_norm(sheet.title)) or _classify_by_content(rows)
        if kind is not None and kind not in found:
            found[kind] = rows
        else:
            leftovers.append((sheet.title, rows))

    # Не опознанное по содержимому раскладываем по похожести названия
    for title, rows in leftovers:
        lowered = _norm(title)
        for kind, hints in _SHEET_HINTS.items():
            if kind not in found and any(hint in lowered for hint in hints):
                found[kind] = rows
                break

    return found


def _find_header(
    rows: Sequence[Sequence[Any]], hints: dict[str, tuple[str, ...]]
) -> tuple[int, dict[str, int]]:
    """Найти строку заголовков: сверху книги бывает подпись или пояснение."""
    best_index, best_mapping, best_score = 0, {}, -1
    for index, row in enumerate(rows[:15]):
        mapping = _detect(row, hints)
        score = len(mapping)
        # Пара «бумага + количество» в подписи или пояснении встретиться не
        # может, поэтому такое сочетание сразу выдаёт настоящую шапку
        if _has_security(mapping) and "quantity" in mapping:
            score += 5
        if score > best_score:
            best_index, best_mapping, best_score = index, mapping, score
    return best_index, best_mapping


def _has_security(mapping: dict[str, int]) -> bool:
    """Есть ли чем опознать бумагу.

    Кода может не быть вовсе: в выписках депозитария бумагу часто задают
    одним ISIN, и требовать колонку с тикером — значит отказаться читать
    такой файл без причины.
    """
    return "secid" in mapping or "isin" in mapping


def _detect(row: Sequence[Any], hints: dict[str, tuple[str, ...]]) -> dict[str, int]:
    """Сопоставить одну строку заголовков с полями по подсказкам."""
    mapping: dict[str, int] = {}
    used: set[int] = set()
    for field, field_hints in hints.items():
        best_index, best_length = None, 0
        for index, cell in enumerate(row):
            if index in used:
                continue
            lowered = _norm(cell)
            if not lowered:
                continue
            for hint in field_hints:
                if hint in lowered and len(hint) > best_length:
                    best_index, best_length = index, len(hint)
        if best_index is not None:
            mapping[field] = best_index
            used.add(best_index)
    return mapping


def _resolve_secid(
    raw_secid: str, raw_isin: str, known: set[str], by_isin: dict[str, str]
) -> str:
    """Найти бумагу по коду или ISIN — в файлах бывает и то, и другое."""
    secid = raw_secid.strip().upper()
    isin = raw_isin.strip().upper()
    if secid in known:
        return secid
    if secid in by_isin:
        return by_isin[secid]
    if isin in by_isin:
        return by_isin[isin]
    return secid


# ----------------------------------------------------------------------
# Предпросмотр
# ----------------------------------------------------------------------
def preview(
    session: Session, content: bytes, filename: str
) -> dict[str, Any]:
    """Разобрать книгу и показать, что будет записано, вместе с проблемами."""
    sheets = read_sheets(content, filename)
    if not sheets:
        return {
            "portfolios": [],
            "holdings": [],
            "deals": [],
            "errors": ["В книге нет ни одного заполненного листа"],
            "warnings": [],
        }

    known = {row[0] for row in session.execute(select(Instrument.secid).distinct()).all()}
    by_isin = {
        isin: secid
        for isin, secid in session.execute(
            select(Instrument.isin, Instrument.secid).where(Instrument.isin.isnot(None))
        ).all()
    }

    portfolios = _read_portfolios(sheets.get(SHEET_PORTFOLIOS, []))
    holdings = _read_holdings(sheets.get(SHEET_HOLDINGS, []), known, by_isin)
    deals = _read_deals(sheets.get(SHEET_DEALS, []), known, by_isin)

    warnings: list[str] = []

    # Портфель, названный в остатках или сделках, но не описанный на листе
    # «Портфели», будет заведён как торговый — предупреждаем, потому что для
    # инвестиционного портфеля это меняет всю переоценку
    described = {item["name"] for item in portfolios}
    mentioned = {
        item["portfolio"]
        for item in holdings + deals
        if item.get("portfolio")
    }
    undescribed = sorted(mentioned - described)
    if undescribed:
        warnings.append(
            "Вид учёта не указан, считаем торговыми: " + ", ".join(undescribed)
        )

    # Главный риск этого импорта — задвоение: остатки уже включают в себя
    # результат прошлых сделок, и если по портфелю сделки уже загружены,
    # позиция удвоится
    existing = {
        row[0]
        for row in session.execute(select(Deal.portfolio).distinct()).all()
        if row[0]
    }
    clashes = sorted({item["portfolio"] for item in holdings} & existing)
    if clashes:
        warnings.append(
            "По этим портфелям сделки уже есть, остатки добавятся сверху и "
            "позиция задвоится: " + ", ".join(clashes)
            + ". Включите «заменить портфель целиком», если файл — это полный текущий состав."
        )

    return {
        "portfolios": portfolios,
        "holdings": holdings,
        "deals": deals,
        "sheets": sorted(sheets),
        "errors": [],
        "warnings": warnings,
        "holdings_valid": sum(1 for item in holdings if item["ok"]),
        "deals_valid": sum(1 for item in deals if item["ok"]),
    }


def _read_portfolios(rows: Sequence[Sequence[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    hints = {
        "name": ("портфель", "название", "наименование", "счет", "счёт"),
        "accounting_type": ("вид учет", "вид учёт", "тип учет", "тип учёт", "учет", "учёт", "категор"),
        "comment": ("коммент", "примеч", "note"),
    }
    header_index, mapping = _find_header(rows, hints)
    if "name" not in mapping:
        return []

    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows[header_index + 1 :]:
        name = str(_cell(row, mapping, "name") or "").strip()
        if not name or name.lower() in seen:
            continue
        seen.add(name.lower())
        accounting = parse_accounting_type(_cell(row, mapping, "accounting_type"))
        items.append(
            {
                "name": name,
                "accounting_type": accounting,
                "accounting_title": ACCOUNTING_TITLES[accounting],
                "comment": str(_cell(row, mapping, "comment") or "").strip() or None,
            }
        )
    return items


def _cell(row: Sequence[Any], mapping: dict[str, int], field: str) -> Any:
    index = mapping.get(field)
    return row[index] if index is not None and index < len(row) else None


def _read_holdings(
    rows: Sequence[Sequence[Any]], known: set[str], by_isin: dict[str, str]
) -> list[dict[str, Any]]:
    if not rows:
        return []
    header_index, mapping = _find_header(rows, _HOLDING_HINTS)
    if not _has_security(mapping) or "quantity" not in mapping:
        return []

    items: list[dict[str, Any]] = []
    for line, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(str(value or "").strip() for value in row):
            continue
        if len(items) >= MAX_ROWS:
            break

        secid = _resolve_secid(
            str(_cell(row, mapping, "secid") or ""),
            str(_cell(row, mapping, "isin") or ""),
            known,
            by_isin,
        )
        quantity = parse_number(_cell(row, mapping, "quantity"))
        price = parse_number(_cell(row, mapping, "price"))

        problems: list[str] = []
        if not secid:
            problems.append("не указана бумага")
        elif secid not in known:
            problems.append(f"инструмент {secid} не найден в справочнике")
        if not quantity:
            problems.append("не разобрано количество")
        if price is None:
            problems.append("не разобрана цена приобретения")

        items.append(
            {
                "line": line,
                "portfolio": str(_cell(row, mapping, "portfolio") or "").strip() or "Основной",
                "secid": secid,
                "quantity": abs(quantity) if quantity else None,
                "price": price,
                "accrued_interest": parse_number(_cell(row, mapping, "accrued_interest")) or 0.0,
                "trade_date": parse_date(_cell(row, mapping, "trade_date")) or date.today(),
                "comment": str(_cell(row, mapping, "comment") or "").strip() or None,
                "problems": problems,
                "ok": not problems,
            }
        )
    return items


def _read_deals(
    rows: Sequence[Sequence[Any]], known: set[str], by_isin: dict[str, str]
) -> list[dict[str, Any]]:
    if not rows:
        return []
    hints = {
        **_HOLDING_HINTS,
        "side": ("направлен", "сторона", "операц", "тип сделки", "side", "b/s"),
        "fee": ("комисс", "сбор", "fee", "издержк"),
        "counterparty": ("контрагент", "брокер", "counterparty"),
    }
    header_index, mapping = _find_header(rows, hints)
    if not _has_security(mapping) or "quantity" not in mapping:
        return []

    items: list[dict[str, Any]] = []
    for line, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(str(value or "").strip() for value in row):
            continue
        if len(items) >= MAX_ROWS:
            break

        secid = _resolve_secid(
            str(_cell(row, mapping, "secid") or ""),
            str(_cell(row, mapping, "isin") or ""),
            known,
            by_isin,
        )
        quantity = parse_number(_cell(row, mapping, "quantity"))
        price = parse_number(_cell(row, mapping, "price"))
        side = parse_side(_cell(row, mapping, "side"))
        # Сторону часто не пишут словом, а кодируют знаком количества
        if side is None and quantity is not None:
            side = "sell" if quantity < 0 else "buy"

        problems: list[str] = []
        if not secid:
            problems.append("не указана бумага")
        elif secid not in known:
            problems.append(f"инструмент {secid} не найден в справочнике")
        if not quantity:
            problems.append("не разобрано количество")
        if price is None:
            problems.append("не разобрана цена")

        items.append(
            {
                "line": line,
                "portfolio": str(_cell(row, mapping, "portfolio") or "").strip() or "Основной",
                "secid": secid,
                "side": side or "buy",
                "quantity": abs(quantity) if quantity else None,
                "price": price,
                "accrued_interest": parse_number(_cell(row, mapping, "accrued_interest")) or 0.0,
                "fee": parse_number(_cell(row, mapping, "fee")) or 0.0,
                "trade_date": parse_date(_cell(row, mapping, "trade_date")) or date.today(),
                "counterparty": str(_cell(row, mapping, "counterparty") or "").strip() or None,
                "problems": problems,
                "ok": not problems,
            }
        )
    return items


# ----------------------------------------------------------------------
# Запись
# ----------------------------------------------------------------------
def apply(
    session: Session,
    payload: dict[str, Any],
    *,
    replace_existing: bool = False,
) -> dict[str, Any]:
    """Записать разобранную книгу.

    ``replace_existing`` стирает прежние сделки затронутых портфелей: это режим
    «файл — полный текущий состав», без него повторная загрузка тех же остатков
    удвоила бы позицию. Портфели, которых в файле нет, не трогаем никогда.
    """
    known = {row[0] for row in session.execute(select(Instrument.secid).distinct()).all()}

    portfolios = payload.get("portfolios") or []
    holdings = payload.get("holdings") or []
    deals = payload.get("deals") or []

    # 1. Справочник портфелей
    saved_portfolios = 0
    for item in portfolios:
        name = str(item.get("name") or "").strip()
        if not name:
            continue
        accounting = parse_accounting_type(item.get("accounting_type"))
        record = session.execute(
            select(Portfolio).where(Portfolio.name == name)
        ).scalar_one_or_none()
        if record is None:
            session.add(
                Portfolio(
                    name=name,
                    accounting_type=accounting,
                    comment=item.get("comment"),
                )
            )
        else:
            record.accounting_type = accounting
            if item.get("comment"):
                record.comment = item["comment"]
        saved_portfolios += 1

    # 2. Замена состава, если запрошена
    touched = {
        str(item.get("portfolio") or "").strip() or "Основной"
        for item in holdings + deals
    }
    removed = 0
    if replace_existing and touched:
        removed = session.execute(
            delete(Deal).where(Deal.portfolio.in_(sorted(touched)))
        ).rowcount or 0

    created_holdings, created_deals = 0, 0
    skipped: list[dict[str, Any]] = []

    def _write(item: dict[str, Any], *, side: str, comment: str) -> bool:
        secid = str(item.get("secid") or "").upper()
        quantity = parse_number(item.get("quantity"))
        price = parse_number(item.get("price"))
        if secid not in known:
            skipped.append({"secid": secid, "detail": "инструмент не найден в справочнике"})
            return False
        if not quantity or price is None:
            skipped.append({"secid": secid, "detail": "не разобраны количество или цена"})
            return False

        session.add(
            Deal(
                portfolio=str(item.get("portfolio") or "").strip() or "Основной",
                secid=secid,
                side=side,
                quantity=abs(quantity),
                price=price,
                accrued_interest=parse_number(item.get("accrued_interest")) or 0.0,
                fee=parse_number(item.get("fee")) or 0.0,
                trade_date=parse_date(item.get("trade_date")) or date.today(),
                counterparty=item.get("counterparty") or None,
                comment=comment,
            )
        )
        return True

    # 3. Остатки — входящими покупками
    for item in holdings:
        if _write(item, side="buy", comment="Входящий остаток"):
            created_holdings += 1

    # 4. Сделки — как есть
    for item in deals:
        side = parse_side(item.get("side")) or "buy"
        if _write(item, side=side, comment="Импорт из файла"):
            created_deals += 1

    session.commit()
    return {
        "portfolios": saved_portfolios,
        "holdings": created_holdings,
        "deals": created_deals,
        "removed": removed,
        "skipped": skipped,
        "skipped_count": len(skipped),
    }
