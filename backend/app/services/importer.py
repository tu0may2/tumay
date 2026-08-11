"""Импорт журнала сделок и сверка позиций с выпиской.

У казначея портфель уже существует — в отчёте брокера или в Excel. Перебивать
его руками невыносимо, поэтому файл распознаётся автоматически: колонки
угадываются по заголовкам, даты и числа приводятся к нужному виду, а перед
записью показывается предпросмотр с найденными ошибками.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Deal, Instrument
from .portfolio import compute_positions

#: Как называют колонки в отчётах брокеров и в самодельных таблицах.
#: Ключ — наше поле, значения — куски заголовков в нижнем регистре.
COLUMN_HINTS: dict[str, tuple[str, ...]] = {
    "secid": ("код", "тикер", "secid", "инструмент", "бумага", "symbol", "ticker"),
    "isin": ("isin", "исин"),
    "side": ("направлен", "сторона", "операц", "тип сделки", "вид", "side", "b/s"),
    "quantity": (
        "количество", "кол-во", "объем, шт", "объём, шт", "qty", "quantity", "лот",
        # Выписки депозитариев пишут не «количество», а «остаток» или «штук»
        "остаток", "штук", "шт.",
    ),
    "price": ("цена", "курс сделки", "price"),
    "accrued_interest": ("нкд", "накопленн", "accrued"),
    "fee": ("комисс", "сбор", "fee", "издержк"),
    "trade_date": ("дата", "date", "дата сделки", "дата заключ"),
    "portfolio": ("портфель", "счет", "счёт", "portfolio", "account"),
    "counterparty": ("контрагент", "брокер", "counterparty"),
}

#: Слова, по которым распознаём сторону сделки
BUY_WORDS = ("покупка", "куп", "buy", "b", "приобрет", "+")
SELL_WORDS = ("продажа", "прод", "sell", "s", "реализ", "-")

#: Максимум строк за один импорт
MAX_ROWS = 5000


def _norm(value: Any) -> str:
    return str(value or "").strip().lower()


def detect_columns(headers: Sequence[str]) -> dict[str, int]:
    """Сопоставить заголовки файла с нашими полями.

    Ищем по вхождению подсказки в заголовок: «Дата сделки» и «дата» одинаково
    попадут в ``trade_date``. Более длинная подсказка важнее короткой, поэтому
    сортируем — иначе «объем, шт» проиграет «объем».
    """
    mapping: dict[str, int] = {}
    used: set[int] = set()

    for field, hints in COLUMN_HINTS.items():
        best_index = None
        best_length = 0
        for index, header in enumerate(headers):
            if index in used:
                continue
            lowered = _norm(header)
            if not lowered:
                continue
            for hint in hints:
                if hint in lowered and len(hint) > best_length:
                    best_index, best_length = index, len(hint)
        if best_index is not None:
            mapping[field] = best_index
            used.add(best_index)
    return mapping


def parse_number(value: Any) -> float | None:
    """Число из ячейки: пробелы-разделители и десятичная запятая допустимы."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-+eE]", "", text)
    if not text or text in ("-", "+", "."):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    # Дата может прийти с временем — берём первую часть
    text = text.split(" ")[0].split("T")[0]
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_side(value: Any) -> str | None:
    lowered = _norm(value)
    if not lowered:
        return None
    for word in BUY_WORDS:
        if lowered.startswith(word):
            return "buy"
    for word in SELL_WORDS:
        if lowered.startswith(word):
            return "sell"
    return None


def read_table(content: bytes, filename: str) -> list[list[Any]]:
    """Прочитать файл как таблицу: поддержаны .xlsx и текстовые форматы."""
    name = (filename or "").lower()

    if name.endswith((".xlsx", ".xlsm", ".xltx")):
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    # Текстовые файлы: кодировка и разделитель у брокеров разные
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Не удалось определить кодировку файла")

    sample = "\n".join(text.splitlines()[:5])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _header_score(mapping: dict[str, int]) -> int:
    """Насколько строка похожа на шапку таблицы.

    Считать одни лишь совпадения мало: в заголовке отчёта «Выписка по счёту
    депо…» слово «счёт» само по себе даёт одно совпадение и перебивает
    настоящую шапку, если в ней распозналась одна колонка. Поэтому за пару
    «бумага + количество или цена» даём отдельную прибавку — такое сочетание
    в подписи встретиться не может.
    """
    score = len(mapping)
    if "secid" in mapping and ("quantity" in mapping or "price" in mapping):
        score += 5
    return score


def _find_header(rows: Sequence[Sequence[Any]]) -> tuple[int, dict[str, int]]:
    """Найти строку заголовков.

    В отчётах брокеров сверху бывает шапка с реквизитами, поэтому проверяем
    первые строки и берём самую похожую на настоящую шапку.
    """
    best_index, best_mapping, best_score = 0, {}, -1
    for index, row in enumerate(rows[:25]):
        mapping = detect_columns([str(cell or "") for cell in row])
        score = _header_score(mapping)
        if score > best_score:
            best_index, best_mapping, best_score = index, mapping, score
    return best_index, best_mapping


def preview_import(
    session: Session,
    content: bytes,
    filename: str,
    *,
    portfolio: str | None = None,
) -> dict[str, Any]:
    """Разобрать файл и показать, что будет загружено."""
    rows = read_table(content, filename)
    if not rows:
        return {"columns": {}, "deals": [], "errors": ["Файл пуст"], "total": 0}

    header_index, mapping = _find_header(rows)
    required = {"secid", "quantity", "price"}
    missing = required - set(mapping)
    if missing:
        titles = {"secid": "бумага", "quantity": "количество", "price": "цена"}
        return {
            "columns": mapping,
            "header_row": header_index,
            "deals": [],
            "errors": [
                "Не найдены обязательные колонки: "
                + ", ".join(titles[field] for field in sorted(missing))
            ],
            "total": 0,
        }

    known = {
        row[0]
        for row in session.execute(select(Instrument.secid).distinct()).all()
    }
    by_isin = {
        isin: secid
        for isin, secid in session.execute(
            select(Instrument.isin, Instrument.secid).where(Instrument.isin.isnot(None))
        ).all()
    }

    deals: list[dict[str, Any]] = []
    errors: list[str] = []

    def cell(row: Sequence[Any], field: str) -> Any:
        index = mapping.get(field)
        return row[index] if index is not None and index < len(row) else None

    for line_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(str(value or "").strip() for value in row):
            continue
        if len(deals) >= MAX_ROWS:
            errors.append(f"Файл обрезан: загружаем не более {MAX_ROWS} строк")
            break

        secid = str(cell(row, "secid") or "").strip().upper()
        isin = str(cell(row, "isin") or "").strip().upper()
        # Если в колонке кода оказался ISIN, находим по нему тикер
        if secid not in known and secid in by_isin:
            secid = by_isin[secid]
        elif secid not in known and isin in by_isin:
            secid = by_isin[isin]

        quantity = parse_number(cell(row, "quantity"))
        price = parse_number(cell(row, "price"))
        trade_date = parse_date(cell(row, "trade_date")) or date.today()
        side = parse_side(cell(row, "side"))

        problems: list[str] = []
        if not secid:
            problems.append("не указана бумага")
        elif secid not in known:
            problems.append(f"инструмент {secid} не найден в справочнике")
        if not quantity:
            problems.append("не разобрано количество")
        if not price:
            problems.append("не разобрана цена")

        # Сторону часто не пишут, а кодируют знаком количества
        if side is None and quantity is not None:
            side = "sell" if quantity < 0 else "buy"

        deals.append(
            {
                "line": line_number,
                "secid": secid,
                "side": side or "buy",
                "quantity": abs(quantity) if quantity else None,
                "price": price,
                "accrued_interest": parse_number(cell(row, "accrued_interest")) or 0.0,
                "fee": parse_number(cell(row, "fee")) or 0.0,
                "trade_date": trade_date,
                "portfolio": (
                    portfolio
                    or str(cell(row, "portfolio") or "").strip()
                    or "Основной"
                ),
                "counterparty": str(cell(row, "counterparty") or "").strip() or None,
                "problems": problems,
                "ok": not problems,
            }
        )

    valid = [deal for deal in deals if deal["ok"]]
    return {
        "columns": mapping,
        "header_row": header_index,
        "headers": [str(cell or "") for cell in rows[header_index]],
        "deals": deals,
        "errors": errors,
        "total": len(deals),
        "valid": len(valid),
        "invalid": len(deals) - len(valid),
    }


def apply_import(session: Session, deals: Sequence[dict[str, Any]]) -> dict[str, Any]:
    """Записать разобранные сделки в журнал.

    Строки приходят обратно из предпросмотра через JSON, поэтому дата в них
    уже текст, а числа могут оказаться строками: приводим типы повторно.
    """
    known = {
        row[0] for row in session.execute(select(Instrument.secid).distinct()).all()
    }

    created = 0
    skipped: list[dict[str, Any]] = []
    for item in deals:
        secid = str(item.get("secid") or "").upper()
        quantity = parse_number(item.get("quantity"))
        price = parse_number(item.get("price"))
        if secid not in known:
            skipped.append({"secid": secid, "detail": "инструмент не найден в справочнике"})
            continue
        if not quantity or price is None:
            skipped.append({"secid": secid, "detail": "не разобраны количество или цена"})
            continue

        side = parse_side(item.get("side")) or ("sell" if quantity < 0 else "buy")

        session.add(
            Deal(
                secid=secid,
                side=side,
                quantity=abs(quantity),
                price=price,
                accrued_interest=parse_number(item.get("accrued_interest")) or 0.0,
                fee=parse_number(item.get("fee")) or 0.0,
                trade_date=parse_date(item.get("trade_date")) or date.today(),
                portfolio=str(item.get("portfolio") or "").strip() or "Основной",
                counterparty=(str(item.get("counterparty")).strip() or None)
                if item.get("counterparty")
                else None,
                comment="Импорт из файла",
            )
        )
        created += 1

    session.commit()
    return {"created": created, "skipped": skipped, "skipped_count": len(skipped)}


# ----------------------------------------------------------------------
# Сверка с выпиской
# ----------------------------------------------------------------------
def reconcile(
    session: Session,
    content: bytes,
    filename: str,
    *,
    portfolio: str | None = None,
) -> dict[str, Any]:
    """Сопоставить позиции терминала с выпиской депозитария.

    Ожидается таблица с кодом бумаги и количеством: остальное не требуется.
    """
    rows = read_table(content, filename)
    if not rows:
        return {"matched": [], "differences": [], "errors": ["Файл пуст"]}

    header_index, mapping = _find_header(rows)
    if "secid" not in mapping or "quantity" not in mapping:
        return {
            "matched": [],
            "differences": [],
            "errors": ["В файле не найдены колонки с бумагой и количеством"],
        }

    by_isin = {
        isin: secid
        for isin, secid in session.execute(
            select(Instrument.isin, Instrument.secid).where(Instrument.isin.isnot(None))
        ).all()
    }

    statement: dict[str, float] = {}
    for row in rows[header_index + 1 :]:
        secid_index, qty_index = mapping["secid"], mapping["quantity"]
        if secid_index >= len(row) or qty_index >= len(row):
            continue
        secid = str(row[secid_index] or "").strip().upper()
        if secid in by_isin:
            secid = by_isin[secid]
        quantity = parse_number(row[qty_index])
        if not secid or quantity is None:
            continue
        statement[secid] = statement.get(secid, 0) + quantity

    positions = {
        p["secid"]: p
        for p in compute_positions(session, portfolio=portfolio)
        if p["quantity"] > 0
    }

    matched: list[dict[str, Any]] = []
    differences: list[dict[str, Any]] = []
    for secid in sorted(set(statement) | set(positions)):
        ours = positions.get(secid, {}).get("quantity", 0.0)
        theirs = statement.get(secid, 0.0)
        delta = round(theirs - ours, 6)
        row = {
            "secid": secid,
            "name": positions.get(secid, {}).get("name", secid),
            "terminal_quantity": ours,
            "statement_quantity": theirs,
            "difference": delta,
            "status": (
                "совпадает" if abs(delta) < 1e-6
                else "нет в терминале" if ours == 0
                else "нет в выписке" if theirs == 0
                else "расхождение"
            ),
        }
        (matched if abs(delta) < 1e-6 else differences).append(row)

    return {
        "matched": matched,
        "differences": differences,
        "matched_count": len(matched),
        "difference_count": len(differences),
        "errors": [],
    }
