"""Платёжный календарь банка: статьи по строкам, дни по столбцам.

Казначейство ведёт календарь не списком событий, а матрицей: слева статьи
(«МБК_тело», «Зарплата», «Налоги»), сверху дни, на пересечении — сумма. Такой
вид отвечает на вопрос «чем закрываем четверг», на который список движений,
отсортированный по дате, отвечает плохо.

Раскладка повторяет рабочий файл казначейства строка в строку: те же
заголовки разделов, те же пустые строки между ними, тот же порядок и та же
заливка. Сохранены и неровности: подряд идут две статьи «5.2», «5. Депозит
ЦБ» стоит выше «5.1. Депозит ЦБ_тело», а «7. Прочее» встречается и в
поступлениях, и в платежах под разными номерами. Переименовывать и
переставлять их здесь нельзя — календарь сверяют глазами с прежним файлом,
и любая «поправленная» строка превращает сверку в поиск отличий.

Пустые строки — тоже часть раскладки, а не мусор: по ним глаз находит
границу блока, и без них длинный столбец статей читается сплошняком.

Суммы попадают в статьи из выгрузки по лицевым счетам: номер счёта определяет
статью, оборот по дебету — платёж, оборот по кредиту — поступление. Правило
разноски применяется при чтении, а не при загрузке, поэтому исправленный
классификатор чинит сразу все загруженные дни.
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, timedelta
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import CalendarEntry, LedgerRow

#: Направление оборота: дебет счёта — деньги ушли, кредит — пришли
DEBIT = "debit"
CREDIT = "credit"

#: Разделы календаря
SECTION_OPENING = "opening"
SECTION_INFLOW = "inflow"
SECTION_OUTFLOW = "outflow"
SECTION_BALANCE = "balance"
SECTION_CLIENT = "client"
SECTION_RATIOS = "ratios"

#: Как строка календаря получает значение:
#: ``ledger`` — из выгрузки, ``computed`` — считается терминалом,
#: ``manual`` — заполняется руками (в выгрузке таких данных нет)
FILL_LEDGER = "ledger"
FILL_COMPUTED = "computed"
FILL_MANUAL = "manual"

#: Что за строка в раскладке файла:
#: ``article`` — статья с суммами, ``caption`` — заголовок раздела,
#: ``spacer`` — пустая строка-разделитель
KIND_ARTICLE = "article"
KIND_CAPTION = "caption"
KIND_SPACER = "spacer"


class Row:
    """Строка календаря: код для расчётов, заголовок — для глаз.

    ``accent`` и ``running`` повторяют заливку рабочего файла: голубым в нём
    отмечены статьи, которые тянутся из системы, зелёным — накопительное
    сальдо. Цвет здесь не украшение: по нему казначей за секунду находит
    строку, которую ищет, и терять его при переносе в терминал нельзя.
    """

    __slots__ = ("code", "title", "section", "fill", "hint", "kind", "accent", "running")

    def __init__(
        self,
        code: str,
        title: str,
        section: str,
        fill: str,
        hint: str = "",
        *,
        kind: str = KIND_ARTICLE,
        accent: bool = False,
        running: bool = False,
    ) -> None:
        self.code = code
        self.title = title
        self.section = section
        self.fill = fill
        self.hint = hint
        self.kind = kind
        self.accent = accent
        self.running = running

    def as_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "title": self.title,
            "section": self.section,
            "fill": self.fill,
            "hint": self.hint,
            "kind": self.kind,
            "accent": self.accent,
            "running": self.running,
        }


def _caption(code: str, title: str, section: str) -> Row:
    return Row(code, title, section, FILL_MANUAL, kind=KIND_CAPTION)


def _spacer(code: str, section: str) -> Row:
    return Row(code, "", section, FILL_MANUAL, kind=KIND_SPACER)


#: Строки календаря ровно в том порядке и с той заливкой, в каких они идут в
#: рабочем файле казначейства (строки 2–55 листа «платежный календарь»)
ROWS: tuple[Row, ...] = (
    Row("opening", "На начало дня Остаток на кор. счете 30102",
        SECTION_OPENING, FILL_LEDGER, "Входящий остаток счёта 30102", accent=True),

    _caption("cap_inflow", "Поступления денежных средств", SECTION_INFLOW),
    Row("in_loans_retail", "1. Кредиты ФЛ", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_loans_corp", "1.1. Кредиты ЮЛ", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_loans_cession", "1.2. Кредиты ФЛ Цессия", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_ibl_principal", "2.1. МБК_тело", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_ibl_interest", "2.2. МБК_ проценты", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_sfuk_principal", "3.1. СФУК_тело", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_sfuk_interest", "3.2. СФУК_проценты", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_deposit_corp", "4. Депозит ЮЛ_тело", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_deposit_cbr", "5. Депозит ЦБ", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_deposit_expo", "5.2. Депозит в Экспо", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_deposit_cbr_principal", "5.1. Депозит ЦБ_тело", SECTION_INFLOW, FILL_LEDGER,
        accent=True),
    Row("in_deposit_cbr_interest", "5.2. Депозит ЦБ_проценты", SECTION_INFLOW, FILL_LEDGER,
        accent=True),
    Row("in_client_accounts", "6. Движения по расчетным счетам клиентов (Поступления)",
        SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_other", "7. Прочее", SECTION_INFLOW, FILL_LEDGER, accent=True),
    Row("in_total", "ИТОГО поступлений", SECTION_INFLOW, FILL_COMPUTED,
        "Сумма статей поступлений", accent=True),

    _spacer("gap_inflow", SECTION_INFLOW),
    _caption("cap_outflow", "Планируемые Платежи", SECTION_OUTFLOW),
    Row("out_loans_cession", "1. Кредиты ФЛ Цессия", SECTION_OUTFLOW, FILL_LEDGER,
        accent=True),
    Row("out_loans_corp", "1.1. Кредиты ЮЛ", SECTION_OUTFLOW, FILL_LEDGER, accent=True),
    Row("out_ibl_principal", "2.1. МБК_тело", SECTION_OUTFLOW, FILL_LEDGER, accent=True),
    Row("out_ibl_interest", "2.2. МБК_проценты", SECTION_OUTFLOW, FILL_LEDGER, accent=True),
    Row("out_sfuk_principal", "3. СФУК_тело", SECTION_OUTFLOW, FILL_LEDGER, accent=True),
    Row("out_deposit_corp_principal", "4.1. Депозит ЮЛ _тело", SECTION_OUTFLOW, FILL_LEDGER,
        accent=True),
    Row("out_deposit_corp_interest", "4.2. Депозиты ЮЛ_проценты", SECTION_OUTFLOW,
        FILL_LEDGER, accent=True),
    Row("out_deposit_cbr", "5.1. Депозит ЦБ", SECTION_OUTFLOW, FILL_LEDGER, accent=True),
    Row("out_deposit_expo", "5.2. Депозит в Экспо", SECTION_OUTFLOW, FILL_LEDGER,
        accent=True),
    Row("out_client_accounts", "6. Движения по расчетным счетам клиентов (Списания)",
        SECTION_OUTFLOW, FILL_LEDGER, accent=True),
    # Строки 31–35 файла заливки не имеют: их казначейство ведёт руками
    Row("out_salary", "7. Зарплата", SECTION_OUTFLOW, FILL_LEDGER),
    Row("out_overhead", "8. Платежи общехозяйственные (поставщикам услуг через WSS)",
        SECTION_OUTFLOW, FILL_LEDGER),
    Row("out_taxes", "9. Налоги", SECTION_OUTFLOW, FILL_LEDGER),
    Row("out_taxes_salary", "10. Налоги зп", SECTION_OUTFLOW, FILL_LEDGER),
    Row("out_other", "11. Прочее", SECTION_OUTFLOW, FILL_LEDGER),
    Row("out_total", "ИТОГО платежей", SECTION_OUTFLOW, FILL_COMPUTED,
        "Сумма статей платежей", accent=True),

    _spacer("gap_outflow", SECTION_BALANCE),
    Row("day_net", "САЛЬДО ДНЯ", SECTION_BALANCE, FILL_COMPUTED,
        "Поступления минус платежи"),
    Row("reserve_for", "в т.ч. ФОР", SECTION_BALANCE, FILL_LEDGER,
        "Обязательные резервы, счёт 302"),
    Row("cumulative", "Накопительное сальдо", SECTION_BALANCE, FILL_COMPUTED,
        "Остаток на начало плюс сальдо всех дней с начала периода", running=True),

    _spacer("gap_balance", SECTION_CLIENT),
    _caption("cap_client", "Остатки на клиентских счетах", SECTION_CLIENT),
    Row("client_407_408", "Счет 407, Счет 408", SECTION_CLIENT, FILL_LEDGER,
        "Исходящий остаток расчётных счетов клиентов"),
    Row("client_420_421", "Счет 420, Счет 421", SECTION_CLIENT, FILL_LEDGER,
        "Исходящий остаток депозитов"),

    _spacer("gap_client", SECTION_RATIOS),
    Row("h2", "Н2 (триггер не менее 17,5)", SECTION_RATIOS, FILL_MANUAL,
        "Считается на вкладке «Нормативы»"),
    _spacer("gap_h2", SECTION_RATIOS),
    Row("h3", "Н3 (триггер не менее 57,50)", SECTION_RATIOS, FILL_MANUAL,
        "Считается на вкладке «Нормативы»"),
    _spacer("gap_h3", SECTION_RATIOS),
    Row("retail_old", "Для Retail и цессия (стар)", SECTION_RATIOS, FILL_MANUAL),
    Row("tranche_1", "Для 1-го транша (не удалять)", SECTION_RATIOS, FILL_MANUAL),
    Row("tranche_2", "Для 2-го транша (не удалять)", SECTION_RATIOS, FILL_MANUAL),
    Row("long_12m", "Остаток со сроком 12+ мес.", SECTION_RATIOS, FILL_MANUAL),
    Row("capital", "Капитал", SECTION_RATIOS, FILL_MANUAL),
    Row("h4", "Н4", SECTION_RATIOS, FILL_MANUAL, "Считается на вкладке «Нормативы»"),
)

#: Только строки с суммами: заголовки и пустые строки в расчётах не участвуют
ARTICLE_ROWS: tuple[Row, ...] = tuple(
    row for row in ROWS if row.kind == KIND_ARTICLE
)

ROW_BY_CODE: dict[str, Row] = {row.code: row for row in ARTICLE_ROWS}

INFLOW_CODES: tuple[str, ...] = tuple(
    row.code for row in ARTICLE_ROWS
    if row.section == SECTION_INFLOW and row.fill == FILL_LEDGER
)
OUTFLOW_CODES: tuple[str, ...] = tuple(
    row.code for row in ARTICLE_ROWS
    if row.section == SECTION_OUTFLOW and row.fill == FILL_LEDGER
)

#: Строки, которые считает сам терминал: вписать в них число нельзя, иначе
#: «ИТОГО» перестало бы быть суммой своих статей и по календарю нельзя было бы
#: проверить ни одну цифру
COMPUTED_CODES: frozenset[str] = frozenset(
    row.code for row in ARTICLE_ROWS if row.fill == FILL_COMPUTED
)

#: Строки, в которые можно вписать сумму руками
EDITABLE_CODES: frozenset[str] = frozenset(
    row.code for row in ARTICLE_ROWS if row.code not in COMPUTED_CODES
)


# ----------------------------------------------------------------------
# Разноска лицевых счетов по статьям
# ----------------------------------------------------------------------
class Rule:
    """Правило разноски: счёт с таким началом и таким оборотом — эта статья."""

    __slots__ = ("prefix", "direction", "row_code", "note", "confirm")

    def __init__(
        self,
        prefix: str,
        direction: str,
        row_code: str,
        note: str = "",
        *,
        confirm: bool = False,
    ) -> None:
        self.prefix = prefix
        self.direction = direction
        self.row_code = row_code
        self.note = note
        #: Разноска угадана, а не задана казначейством — показываем в разборе
        self.confirm = confirm

    def as_dict(self) -> dict[str, Any]:
        row = ROW_BY_CODE[self.row_code]
        return {
            "prefix": self.prefix,
            "direction": self.direction,
            "direction_title": "дебет" if self.direction == DEBIT else "кредит",
            "row_code": self.row_code,
            "row_title": row.title,
            "section": row.section,
            "note": self.note,
            "confirm": self.confirm,
        }


#: Классификатор счетов, продиктованный казначейством.
#:
#: Правила разбираются по длине префикса: чем длиннее, тем важнее. Поэтому
#: отдельный лицевой счёт зарплаты перебивает общее правило «603 — общехоз»,
#: и добавление нового исключения не требует переписывать остальные строки.
RULES: tuple[Rule, ...] = (
    # --- Дебет: деньги ушли ------------------------------------------
    Rule("60305810400000000001", DEBIT, "out_salary", "перечисления по ЗП"),
    Rule("60301810800000000100", DEBIT, "out_taxes", "единый налоговый платёж"),
    Rule("603", DEBIT, "out_overhead", "общехозяйственные платежи"),
    # 47426 «обязательства по уплате процентов» и 47422 «обязательства по
    # прочим операциям» — через них проходят проценты по депозитам
    Rule("47426", DEBIT, "out_deposit_corp_interest", "проценты по депозиту"),
    Rule("47422", DEBIT, "out_deposit_corp_interest", "проценты по депозиту"),
    Rule("474", DEBIT, "out_other", "прочее; сюда же попадают платежи по СФУК",
         confirm=True),
    Rule("458", DEBIT, "out_other", "прочее; сюда же попадают платежи по СФУК",
         confirm=True),
    Rule("421", DEBIT, "out_deposit_corp_principal", "вывод депозита"),
    Rule("420", DEBIT, "out_deposit_corp_principal", "вывод депозита"),
    Rule("407", DEBIT, "out_client_accounts", "платежи по расчётным счетам клиентов"),
    Rule("408", DEBIT, "out_client_accounts", "платежи по расчётным счетам клиентов"),
    Rule("319", DEBIT, "out_deposit_cbr", "депозит в ЦБ"),
    Rule("320", DEBIT, "out_ibl_principal", "МБК"),

    # --- Кредит: деньги пришли ---------------------------------------
    Rule("458", CREDIT, "in_loans_retail", "входящие платежи по кредитам",
         confirm=True),
    Rule("474", CREDIT, "in_loans_corp", "входящие платежи по кредитам",
         confirm=True),
    Rule("421", CREDIT, "in_deposit_corp", "входящий депозит"),
    Rule("420", CREDIT, "in_deposit_corp", "входящий депозит"),
    Rule("313", CREDIT, "in_deposit_corp", "депозит ЮЛ", confirm=True),
    Rule("30109", CREDIT, "in_client_accounts", "движения по расчётным счетам"),
    Rule("407", CREDIT, "in_client_accounts", "движения по расчётным счетам"),
    Rule("408", CREDIT, "in_client_accounts", "движения по расчётным счетам"),
    Rule("603", CREDIT, "in_other", "прочие поступления"),
    Rule("459", CREDIT, "in_other", "прочие поступления"),
    Rule("306", CREDIT, "in_other", "вывод с брокерского счёта"),
    Rule("319", CREDIT, "in_deposit_cbr_principal", "возврат депозита ЦБ"),
    Rule("320", CREDIT, "in_ibl_principal", "возврат МБК"),
)

#: Правила в порядке убывания длины префикса — так «самое частное побеждает»
#: становится свойством данных, а не заботой каждого места, где идёт поиск
_SORTED_RULES: tuple[Rule, ...] = tuple(
    sorted(RULES, key=lambda rule: len(rule.prefix), reverse=True)
)

#: Счёт корсчёта в Банке России: с него начинается день
CORR_ACCOUNT_PREFIX = "30102"
#: Фонд обязательных резервов
RESERVE_PREFIX = "302"
#: Остатки клиентов, которые календарь показывает отдельными строками
CLIENT_BALANCE_PREFIXES: dict[str, tuple[str, ...]] = {
    "client_407_408": ("407", "408"),
    "client_420_421": ("420", "421"),
}


#: Счета, обороты которых статьёй календаря быть не могут.
#:
#: 30102 — сам корсчёт: его оборот это зеркало всех остальных проводок разом,
#: и разнести его по статьям значит посчитать день дважды. 302 — фонд
#: обязательных резервов, он в календаре отдельная справочная строка.
#: Без этого списка каждая загрузка честно сообщала бы о «неразнесённых
#: миллионах», которых на самом деле нет.
TECHNICAL_PREFIXES: tuple[str, ...] = (CORR_ACCOUNT_PREFIX, RESERVE_PREFIX)


def normalise_account(value: Any) -> str:
    """Номер счёта без пробелов и точек — в выгрузках их ставят по-разному."""
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def is_technical(account: str) -> bool:
    """Счёт, который в статьи календаря не разносится по замыслу."""
    return normalise_account(account).startswith(TECHNICAL_PREFIXES)


def classify(account: str, direction: str) -> Rule | None:
    """Найти статью календаря для счёта и стороны оборота."""
    digits = normalise_account(account)
    if not digits:
        return None
    for rule in _SORTED_RULES:
        if rule.direction == direction and digits.startswith(rule.prefix):
            return rule
    return None


def rules_table() -> list[dict[str, Any]]:
    """Классификатор для показа человеку — им сверяют разноску."""
    return [rule.as_dict() for rule in _SORTED_RULES]


# ----------------------------------------------------------------------
# Матрица
# ----------------------------------------------------------------------
def loaded_dates(session: Session) -> list[date]:
    """Дни, на которые выгрузка уже загружена."""
    rows = session.execute(
        select(LedgerRow.load_date).distinct().order_by(LedgerRow.load_date)
    ).scalars()
    return list(rows)


def manual_entries(
    session: Session,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
) -> dict[date, dict[str, float]]:
    """Суммы, вписанные руками, по дням и статьям."""
    statement = select(CalendarEntry)
    if date_from is not None:
        statement = statement.where(CalendarEntry.entry_date >= date_from)
    if date_to is not None:
        statement = statement.where(CalendarEntry.entry_date <= date_to)

    result: dict[date, dict[str, float]] = defaultdict(dict)
    for entry in session.execute(statement).scalars():
        result[entry.entry_date][entry.row_code] = entry.amount
    return dict(result)


def save_entry(
    session: Session,
    *,
    entry_date: date,
    row_code: str,
    amount: float | None,
    comment: str | None = None,
    author: str | None = None,
) -> dict[str, Any]:
    """Вписать сумму в ячейку календаря. ``None`` стирает ввод.

    Стирание возвращает ячейку к тому, что дала выгрузка, а не обнуляет её:
    пустая ячейка и ноль — разные утверждения, и ввод не должен уметь
    навсегда закрыть собой факт.
    """
    if row_code not in EDITABLE_CODES:
        raise ValueError(
            f"В строку «{ROW_BY_CODE[row_code].title}» вписать нельзя: она считается"
            if row_code in ROW_BY_CODE
            else "Такой строки в календаре нет"
        )

    record = session.execute(
        select(CalendarEntry).where(
            CalendarEntry.entry_date == entry_date,
            CalendarEntry.row_code == row_code,
        )
    ).scalar_one_or_none()

    if amount is None:
        if record is not None:
            session.delete(record)
            session.commit()
        return {"entry_date": entry_date, "row_code": row_code, "amount": None}

    if record is None:
        record = CalendarEntry(entry_date=entry_date, row_code=row_code, amount=0.0)
        session.add(record)
    record.amount = round(float(amount), 2)
    record.comment = comment
    record.author = author
    session.commit()

    return {
        "entry_date": entry_date,
        "row_code": row_code,
        "amount": record.amount,
        "row_title": ROW_BY_CODE[row_code].title,
    }


def _blank_column() -> dict[str, float | None]:
    return {row.code: None for row in ARTICLE_ROWS}


def _fill_day(
    entries: Sequence[LedgerRow], manual: dict[str, float] | None = None
) -> tuple[dict[str, float | None], dict[str, float | None]]:
    """Свести день календаря: выгрузка плюс вписанные руками суммы.

    Возвращает колонку показываемых значений и колонку того, что дала одна
    выгрузка. Вторая нужна, чтобы у ячейки с ручным вводом было видно
    перекрытое значение: иначе вписанная сумма молча прячет факт, прошедший
    по счетам, и разойтись с выпиской можно было бы, ничего не заметив.

    День без выгрузки и без ввода остаётся пустым целиком, а не нулевым: ноль
    в статье означает «оборота не было», и подставлять его там, где данных
    просто нет, значит утверждать то, чего мы не знаем.
    """
    column: dict[str, float | None] = _blank_column()
    manual = manual or {}
    if not entries and not manual:
        return column, _blank_column()

    buckets: dict[str, float] = defaultdict(float)
    opening = 0.0
    reserve = 0.0
    client: dict[str, float] = defaultdict(float)
    seen_opening = False
    seen_reserve = False

    for entry in entries:
        digits = normalise_account(entry.account)

        if digits.startswith(CORR_ACCOUNT_PREFIX):
            opening += entry.opening_balance
            seen_opening = True
        if digits.startswith(RESERVE_PREFIX):
            reserve += entry.closing_balance
            seen_reserve = True
        for code, prefixes in CLIENT_BALANCE_PREFIXES.items():
            if digits.startswith(prefixes):
                # Пассивные счета клиентов ведутся по кредиту, а остаток
                # показываем как сумму обязательств — знак снимаем
                client[code] += abs(entry.closing_balance)

        for direction, turnover in ((DEBIT, entry.debit_turnover),
                                    (CREDIT, entry.credit_turnover)):
            if not turnover:
                continue
            rule = classify(digits, direction)
            if rule is None:
                continue
            buckets[rule.row_code] += abs(turnover)

    for code, value in buckets.items():
        column[code] = round(value, 2)
    if seen_opening:
        column["opening"] = round(opening, 2)
    if seen_reserve:
        column["reserve_for"] = round(reserve, 2)
    for code, value in client.items():
        column[code] = round(value, 2)

    # Что дала одна выгрузка — запоминаем до наложения ручного ввода
    from_ledger = dict(column)

    # Вписанное руками перекрывает вычисленное: последнее слово за человеком,
    # он же видит перекрытую сумму в подсказке и может стереть свой ввод
    for code, value in manual.items():
        if code in EDITABLE_CODES:
            column[code] = round(value, 2)

    inflow = sum(column.get(code) or 0.0 for code in INFLOW_CODES)
    outflow = sum(column.get(code) or 0.0 for code in OUTFLOW_CODES)
    column["in_total"] = round(inflow, 2)
    column["out_total"] = round(outflow, 2)
    column["day_net"] = round(inflow - outflow, 2)
    return column, from_ledger


def matrix(
    session: Session,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    only_loaded: bool = False,
) -> dict[str, Any]:
    """Календарь как таблица: строки — статьи, столбцы — дни.

    ``only_loaded`` показывает лишь дни, на которые есть выгрузка или ручной
    ввод. По умолчанию колонки идут подряд, включая пустые: разрыв в датах —
    это сообщение («за среду выгрузку не загрузили»), а не то, что стоит
    прятать.
    """
    loaded = loaded_dates(session)
    manual = manual_entries(session)

    known = sorted(set(loaded) | set(manual))
    if date_from is None:
        date_from = known[0] if known else date.today()
    if date_to is None:
        date_to = known[-1] if known else date_from

    if date_to < date_from:
        date_from, date_to = date_to, date_from

    if only_loaded:
        days = [day for day in known if date_from <= day <= date_to]
    else:
        span = (date_to - date_from).days
        days = [date_from + timedelta(days=offset) for offset in range(span + 1)]

    entries = session.execute(
        select(LedgerRow).where(
            LedgerRow.load_date >= date_from, LedgerRow.load_date <= date_to
        )
    ).scalars()

    by_day: dict[date, list[LedgerRow]] = defaultdict(list)
    for entry in entries:
        by_day[entry.load_date].append(entry)

    columns: dict[date, dict[str, float | None]] = {}
    ledger_columns: dict[date, dict[str, float | None]] = {}
    for day in days:
        columns[day], ledger_columns[day] = _fill_day(
            by_day.get(day, []), manual.get(day)
        )

    # Накопительное сальдо тянется через все показанные дни: без него не видно,
    # что три спокойных дня подряд всё равно съедают остаток. Идёт по дням,
    # про которые хоть что-то известно, — вписанный вручную план будущего
    # платежа обязан двигать остаток, иначе разрыв не увидеть заранее
    filled = {day for day in days if day in by_day or day in manual}
    running: float | None = None
    for day in days:
        if day not in filled:
            continue
        column = columns[day]
        if running is None:
            # Отсчёт ведём от корсчёта первого известного дня: сальдо само по
            # себе показывает только динамику, а вопрос всегда в том, хватит ли
            # денег — то есть в абсолютном остатке
            opening = column["opening"]
            running = opening if opening is not None else 0.0
        running += column["day_net"] or 0.0
        column["cumulative"] = round(running, 2)

    loaded_set = set(by_day)

    def _row(row: Row) -> dict[str, Any]:
        if row.kind != KIND_ARTICLE:
            # У заголовка и пустой строки сумм нет, но пустые списки отдаём
            # всё равно: тогда клиенту не нужно знать заранее, у каких строк
            # колонки есть, а у каких нет
            return {**row.as_dict(), "values": [], "manual": [], "beneath": []}

        editable = row.code in EDITABLE_CODES
        typed = [
            index for index, day in enumerate(days)
            if editable and row.code in manual.get(day, {})
        ]
        return {
            **row.as_dict(),
            "values": [columns[day][row.code] for day in days],
            "editable": editable,
            #: Номера дней, где сумма вписана руками
            "manual": typed,
            #: Что под ручным вводом дала выгрузка — показываем в подсказке
            "beneath": [ledger_columns[days[index]][row.code] for index in typed],
        }

    return {
        "date_from": date_from,
        "date_to": date_to,
        "days": [
            {
                "date": day,
                "loaded": day in loaded_set,
                "typed": day in manual,
                "weekend": day.weekday() >= 5,
            }
            for day in days
        ],
        "rows": [_row(row) for row in ROWS],
        "loaded_days": len(loaded_set),
        "typed_days": len({day for day in days if day in manual}),
        "empty_days": len(days) - len(filled),
        "loaded_dates": [day for day in loaded if date_from <= day <= date_to],
    }


def ledger_sheet(
    session: Session, *, on_date: date | None = None
) -> dict[str, Any]:
    """Лист «Счета»: выгрузка на дату с проставленными статьями календаря."""
    loaded = loaded_dates(session)
    if on_date is None:
        on_date = loaded[-1] if loaded else date.today()

    entries = list(
        session.execute(
            select(LedgerRow)
            .where(LedgerRow.load_date == on_date)
            .order_by(LedgerRow.account)
        ).scalars()
    )

    rows: list[dict[str, Any]] = []
    unmapped = 0
    for entry in entries:
        debit_rule = classify(entry.account, DEBIT) if entry.debit_turnover else None
        credit_rule = classify(entry.account, CREDIT) if entry.credit_turnover else None
        technical = is_technical(entry.account)
        turnover = bool(entry.debit_turnover or entry.credit_turnover)
        if turnover and not technical and debit_rule is None and credit_rule is None:
            unmapped += 1

        rows.append(
            {
                "account": entry.account,
                "account_name": entry.account_name,
                "currency": entry.currency,
                "opening_balance": round(entry.opening_balance, 2),
                "debit_turnover": round(entry.debit_turnover, 2),
                "credit_turnover": round(entry.credit_turnover, 2),
                "closing_balance": round(entry.closing_balance, 2),
                "debit_row": debit_rule.row_code if debit_rule else None,
                "debit_title": (
                    ROW_BY_CODE[debit_rule.row_code].title if debit_rule else None
                ),
                "credit_row": credit_rule.row_code if credit_rule else None,
                "credit_title": (
                    ROW_BY_CODE[credit_rule.row_code].title if credit_rule else None
                ),
                "confirm": bool(
                    (debit_rule and debit_rule.confirm)
                    or (credit_rule and credit_rule.confirm)
                ),
                "technical": technical,
            }
        )

    return {
        "load_date": on_date,
        "dates": loaded,
        "rows": rows,
        "accounts": len(rows),
        "unmapped": unmapped,
        "debit_total": round(sum(row["debit_turnover"] for row in rows), 2),
        "credit_total": round(sum(row["credit_turnover"] for row in rows), 2),
    }


# ----------------------------------------------------------------------
# Выгрузка в Excel
# ----------------------------------------------------------------------
#: Колонки листа «Счета» при выгрузке
LEDGER_COLUMNS: tuple[dict[str, Any], ...] = (
    {"code": "account", "title": "Лицевой счёт", "kind": "text"},
    {"code": "account_name", "title": "Наименование", "kind": "text"},
    {"code": "currency", "title": "Валюта", "kind": "text"},
    {"code": "opening_balance", "title": "Входящий остаток", "kind": "number", "digits": 2},
    {"code": "debit_turnover", "title": "Оборот по дебету", "kind": "number", "digits": 2},
    {"code": "credit_turnover", "title": "Оборот по кредиту", "kind": "number", "digits": 2},
    {"code": "closing_balance", "title": "Исходящий остаток", "kind": "number", "digits": 2},
    {"code": "debit_title", "title": "Статья (дебет)", "kind": "text"},
    {"code": "credit_title", "title": "Статья (кредит)", "kind": "text"},
)

#: Цвета взяты пипеткой из рабочего файла казначейства: голубым в нём отмечены
#: статьи, которые тянутся из системы, зелёным — накопительное сальдо
_FILL_ACCENT = PatternFill("solid", fgColor="00B0F0")
_FILL_RUNNING = PatternFill("solid", fgColor="92D050")
_FILL_WEEKEND = PatternFill("solid", fgColor="F2F2F2")
_THIN = Side(style="thin", color="000000")
_MEDIUM = Side(style="medium", color="000000")
_MONEY = "# ##0.00"

#: Размер шрифта в файле — 11 для статей и 8 для дат в шапке
_TITLE_FONT = Font(bold=True, size=11)
_PLAIN_FONT = Font(size=11)
_DATE_FONT = Font(bold=True, size=8)
#: Вписанное руками — курсивом, чтобы план не выдавал себя за факт
_TYPED_FONT = Font(size=11, italic=True)

#: Строки, у которых в файле нет рамки: аналитический блок под календарём
#: отделён от него пустой строкой и рамкой не обведён
_NO_BORDER_SECTIONS = {SECTION_RATIOS}

#: В файле жирным набран весь столбец статей, кроме двух строк остатков
#: клиентов — их набрали обычным начертанием
_LIGHT_LABEL_ROWS = {"client_407_408", "client_420_421"}


def build_workbook(
    result: dict[str, Any], ledger: dict[str, Any] | None = None
) -> bytes:
    """Собрать книгу в том же виде, в каком календарь ведут в казначействе.

    Раскладка повторяет исходный файл строка в строку: заголовки разделов,
    пустые строки между блоками, заливка и рамки на своих местах. Это не
    аккуратность ради аккуратности — файл уходит людям, которые годами читают
    его глазами, и перестроенная «как удобнее программе» таблица заставила бы
    их заново искать каждую строку.

    Два отличия от исходника сделаны сознательно. Даты пишем форматом
    ДД.ММ.ГГГГ вместо стоявшего в файле mm-dd-yy: американский порядок в
    русском календаре читается как другая дата, и это ошибка, а не решение.
    Суммам задан денежный формат вместо «Общего» — иначе Excel показывает
    12300000 там, где нужно 12 300 000,00.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "платежный календарь"

    days = result["days"]
    header = sheet.cell(row=1, column=1, value="Статья/Дата")
    header.font = _TITLE_FONT
    header.border = Border(top=_THIN, bottom=_THIN)
    for index, day in enumerate(days, start=2):
        cell = sheet.cell(row=1, column=index, value=day["date"])
        cell.number_format = "DD.MM.YYYY"
        cell.font = _DATE_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=_THIN, bottom=_THIN)
        if day["weekend"]:
            cell.fill = _FILL_WEEKEND

    for offset, row in enumerate(result["rows"], start=2):
        # Пустая строка-разделитель: в файле она есть, и место занимает
        if row["kind"] == KIND_SPACER:
            continue

        bordered = row["section"] not in _NO_BORDER_SECTIONS
        # Заголовок «Остатки на клиентских счетах» в файле обведён жирнее
        side = _MEDIUM if row["kind"] == KIND_CAPTION and bordered else _THIN
        border = Border(top=side, bottom=side) if bordered else Border()

        label = sheet.cell(row=offset, column=1, value=row["title"])
        label.font = _PLAIN_FONT if row["code"] in _LIGHT_LABEL_ROWS else _TITLE_FONT
        label.border = border
        if row["accent"]:
            label.fill = _FILL_ACCENT
        elif row["running"]:
            label.fill = _FILL_RUNNING

        typed = set(row.get("manual") or ())
        for index, value in enumerate(row["values"]):
            cell = sheet.cell(row=offset, column=index + 2)
            if value is not None:
                cell.value = value
                cell.number_format = _MONEY
            # Вписанное руками набираем курсивом: в файле, ушедшем по почте,
            # иначе не отличить план от того, что прошло по счетам
            cell.font = _TYPED_FONT if index in typed else _PLAIN_FONT
            cell.border = border
            if row["running"]:
                cell.fill = _FILL_RUNNING
            elif days[index]["weekend"]:
                cell.fill = _FILL_WEEKEND

    # Ширина столбца статей в файле 29, но при ней длинные названия обрезаны;
    # берём такую, при которой самая длинная статья видна целиком
    sheet.column_dimensions["A"].width = 46
    for index in range(2, len(days) + 2):
        sheet.column_dimensions[get_column_letter(index)].width = 15
    sheet.freeze_panes = "B2"

    if ledger is not None:
        _write_ledger_sheet(workbook.create_sheet("Счета"), ledger)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_ledger_sheet(sheet: Any, ledger: dict[str, Any]) -> None:
    """Второй лист книги: из чего сложились суммы календаря."""
    sheet.cell(row=1, column=1, value="Выгрузка на").font = Font(bold=True)
    date_cell = sheet.cell(row=1, column=2, value=ledger["load_date"])
    date_cell.number_format = "DD.MM.YYYY"

    for index, column in enumerate(LEDGER_COLUMNS, start=1):
        cell = sheet.cell(row=3, column=index, value=column["title"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = (
            max(len(column["title"]) + 2, 14) if column["kind"] == "text" else 18
        )

    for offset, row in enumerate(ledger["rows"], start=4):
        for index, column in enumerate(LEDGER_COLUMNS, start=1):
            value = row.get(column["code"])
            cell = sheet.cell(row=offset, column=index)
            if value is None:
                continue
            if column["kind"] == "number":
                cell.value = value
                cell.number_format = _MONEY
            else:
                # Номер счёта — текст: иначе Excel съест ведущие нули
                cell.value = str(value)

    sheet.freeze_panes = "A4"
