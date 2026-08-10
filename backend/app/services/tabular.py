"""Формирование файлов выгрузки: Excel (.xlsx) и CSV."""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

#: Русский Excel по умолчанию ждёт «;» как разделитель полей
CSV_DELIMITER = ";"
#: BOM нужен, чтобы Excel открыл UTF-8 без крякозябр
CSV_BOM = "﻿"

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _number_format(digits: int) -> str:
    """Формат числа с разделителем разрядов под нужную точность."""
    if digits <= 0:
        return "# ##0"
    return "# ##0." + "0" * digits


def to_xlsx(
    columns: Sequence[dict[str, Any]],
    rows: Sequence[dict[str, Any]],
    *,
    sheet_title: str = "Данные",
    meta: Sequence[tuple[str, str]] = (),
) -> bytes:
    """Собрать книгу Excel: числа числами, даты датами, шапка закреплена."""
    workbook = Workbook()
    sheet = workbook.active
    # Excel не принимает в имени листа : \ / ? * [ ] и длину свыше 31 символа
    sheet.title = sheet_title[:31] or "Данные"

    header_row = 1
    if meta:
        # Параметры запроса пишем сверху, чтобы выгрузка была самодостаточной
        for offset, (label, value) in enumerate(meta):
            sheet.cell(row=offset + 1, column=1, value=label).font = Font(bold=True)
            sheet.cell(row=offset + 1, column=2, value=value)
        header_row = len(meta) + 2

    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column["title"])
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=header_row + 1):
        for col_index, column in enumerate(columns, start=1):
            value = row.get(column["code"])
            cell = sheet.cell(row=row_index, column=col_index)
            if value is None:
                continue
            if column.get("kind") == "date":
                cell.value = value if isinstance(value, (date, datetime)) else str(value)
                cell.number_format = "DD.MM.YYYY"
            elif column.get("kind") == "number" and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = _number_format(int(column.get("digits", 2)))
            else:
                cell.value = str(value)

    # Ширина колонок по содержимому, с разумным потолком
    for index, column in enumerate(columns, start=1):
        longest = len(str(column["title"]))
        for row in rows[:200]:
            value = row.get(column["code"])
            if value is not None:
                longest = max(longest, len(str(value)))
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 42)

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(rows)}"
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_csv(
    columns: Sequence[dict[str, Any]],
    rows: Sequence[dict[str, Any]],
    *,
    decimal_comma: bool = True,
) -> bytes:
    """Собрать CSV, который русский Excel открывает без настройки импорта."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=CSV_DELIMITER, lineterminator="\r\n")
    writer.writerow([column["title"] for column in columns])

    for row in rows:
        line = []
        for column in columns:
            value = row.get(column["code"])
            if value is None:
                line.append("")
            elif column.get("kind") == "date":
                line.append(
                    value.strftime("%d.%m.%Y")
                    if isinstance(value, (date, datetime))
                    else str(value)
                )
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                digits = int(column.get("digits", 2))
                text = f"{value:.{digits}f}"
                # Десятичная запятая — иначе Excel с русской локалью
                # прочитает число как текст
                line.append(text.replace(".", ",") if decimal_comma else text)
            else:
                line.append(str(value))
        writer.writerow(line)

    return (CSV_BOM + buffer.getvalue()).encode("utf-8")
